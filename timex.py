#!/usr/bin/env python3
"""
Timex — time tracker for your terminal.

Usage:
    python timex.py

Commands:
    /start   — Start the timer
    /pause   — Pause the timer
    /resume  — Resume the timer
    /new     — Save the session and start fresh
    /export  — Report a period: .html page or .xlsx
    /clear   — Discard the current session
    <text>   — Log a new task (while timer is running)
"""

from __future__ import annotations

import base64
import io
import json
import logging
import os
import subprocess
import threading
import time as _time
import traceback
import webbrowser
from datetime import date, datetime, timedelta
from html import escape as _esc
from pathlib import Path
import re
import secrets
import shutil
import tempfile
from dataclasses import dataclass

from textual.app import App, ComposeResult
from textual.containers import VerticalScroll
from textual.widgets import Static, Input
from textual.widgets._input import Suggester
from textual.binding import Binding
from textual.events import Key
from textual import on

from rich.text import Text
from rich.panel import Panel
from rich.table import Table
from rich.align import Align
from rich.console import Group


# ── Constants ────────────────────────────────────────────────────────────────

IDLE = "idle"
RUNNING = "running"
PAUSED = "paused"

DEFAULT_ACCENT = "#e8a55d"
DEFAULT_ACCENT_HEX = "E8A55D"

DIM = "#555555"
DIMMER = "#333333"
SEPARATOR = "#222222"
TEXT_COLOR = "#d4d4d4"


DEFAULT_REMINDER_INTERVAL = 30 * 60  # 30 minutes in seconds

STATE_DIR = Path.home() / ".timex"
PROJECTS_DIR = STATE_DIR / "projects"
BACKUP_DIR = STATE_DIR / "backups"
BACKUP_KEEP = 30  # the whole tree is ~200 KB, so keeping a month of them is free
ACTIVE_PROJECT_FILE = STATE_DIR / "active_project"
AUTOSAVE_INTERVAL = 30  # seconds between autosaves during tick
CONFIG_FILE = STATE_DIR / "config.json"
CRASH_LOG = STATE_DIR / "crash.log"

VERSION = "1.3.2"
# Patching these in place rewrites files inside the bundle, which breaks the
# notarised signature. Updates therefore ship as a fresh signed app: changelog
# entries default to dmg_required, and self-patching is opt-in per release.


# ── Cyrillic → Latin map (ЙЦУКЕН → QWERTY keyboard layout) ──────────────────

_CYR2LAT = str.maketrans(
    "йцукенгшщзхъфывапролджэячсмитьбюЙЦУКЕНГШЩЗХЪФЫВАПРОЛДЖЭЯЧСМИТЬБЮ",
    "qwertyuiop[]asdfghjkl;'zxcvbnm,.QWERTYUIOP[]ASDFGHJKL;'ZXCVBNM,.",
)


def _translit_cmd(text: str) -> str:
    """Transliterate cyrillic to latin for command input after /."""
    if text.startswith("/") and len(text) > 1:
        return "/" + text[1:].translate(_CYR2LAT)
    return text


# ── Data ─────────────────────────────────────────────────────────────────────


@dataclass
class TaskEntry:
    """A single tracked task."""

    name: str
    wall_start: datetime
    active_start: float          # active seconds at task start
    active_end: float | None = None   # active seconds at task end
    wall_end: datetime | None = None  # wall clock when task ended

    def get_duration(self, current_active: float | None = None) -> float:
        end = self.active_end if self.active_end is not None else (current_active or self.active_start)
        return max(0, end - self.active_start)

    def format_duration(self, current_active: float | None = None) -> str:
        s = int(self.get_duration(current_active))
        h, remainder = divmod(s, 3600)
        m, sec = divmod(remainder, 60)
        if h > 0:
            return f"{h}h {m:02d}m {sec:02d}s"
        if m > 0:
            return f"{m}m {sec:02d}s"
        return f"{sec}s"

    def format_start(self) -> str:
        return self.wall_start.strftime("%H:%M")


# ── Command suggestions ──────────────────────────────────────────────────────

STATE_COMMANDS: dict[str, list[str]] = {
    IDLE:    ["/start", "/new", "/date", "/stats", "/export", "/edit", "/clear", "/help", "/notification", "/project"],
    RUNNING: ["/pause", "/add", "/remove", "/reset", "/new", "/clear", "/date", "/stats", "/export", "/edit", "/help", "/notification", "/project"],
    PAUSED:  ["/resume", "/add", "/remove", "/reset", "/new", "/clear", "/date", "/stats", "/export", "/edit", "/help", "/notification", "/project"],
}


class CommandSuggester(Suggester):
    """Suggest commands based on current app state."""

    def __init__(self, app_ref: object | None = None) -> None:
        super().__init__(use_cache=False)
        self._app_ref = app_ref

    async def get_suggestion(self, value: str) -> str | None:
        try:
            if not value.startswith("/"):
                return None
            val = value.lower()
            # On non-timeline views, suggest /back first
            view_mode = getattr(self._app_ref, "_view_mode", "timeline") if self._app_ref else "timeline"
            if view_mode != "timeline":
                if "/back".startswith(val) and "/back" != val:
                    return "/back"
                if view_mode == "project" and "/edit".startswith(val) and "/edit" != val:
                    return "/edit"
                return None
            state = getattr(self._app_ref, "state", IDLE) if self._app_ref else IDLE
            commands = STATE_COMMANDS.get(state, STATE_COMMANDS[IDLE])
            for cmd in commands:
                if cmd.startswith(val) and cmd != val:
                    return cmd
            return None
        except Exception:
            return None


# ── Input with history ───────────────────────────────────────────────────────


class HistoryInput(Input):
    """Input widget with shell-like Up/Down history navigation."""

    def __init__(self, app_ref: object | None = None, **kwargs) -> None:
        kwargs.setdefault("suggester", CommandSuggester(app_ref))
        super().__init__(**kwargs)
        self._app_ref = app_ref
        self._history: list[str] = []
        self._history_index: int = -1
        self._draft: str = ""  # saves current input when browsing history

    def add_to_history(self, text: str) -> None:
        if text and (not self._history or self._history[-1] != text):
            self._history.append(text)
        self._history_index = -1
        self._draft = ""

    async def _on_key(self, event: Key) -> None:
        try:
            handled = self._handle_key(event)
            if not handled:
                await super()._on_key(event)
        except Exception:
            try:
                with open(CRASH_LOG, "w") as f:
                    traceback.print_exc(file=f)
            except OSError:
                pass

    def _handle_key(self, event: Key) -> bool:
        """Handle custom keys. Returns True if the event was consumed."""
        # Ctrl+A — select all text (macOS Cmd+A maps to ctrl+a in terminal)
        if event.key == "ctrl+a":
            if self.value:
                self.cursor_position = 0
                self.selection = (0, len(self.value))
            event.prevent_default()
            event.stop()
            return True

        # Ctrl+U — delete entire line (macOS Cmd+Backspace equivalent)
        if event.key == "ctrl+u":
            self.value = ""
            self.cursor_position = 0
            event.prevent_default()
            event.stop()
            return True

        if event.key == "tab":
            # Accept suggestion on Tab
            if self._suggestion:
                self.value = self._suggestion
                self.cursor_position = len(self.value)
                event.prevent_default()
                event.stop()
                return True

        # Project edit mode navigation
        app = self._app_ref
        if app and getattr(app, "_view_mode", "") == "project_edit" and getattr(app, "_project_editing", None) is None:
            if event.key == "up":
                event.prevent_default()
                event.stop()
                app._project_edit_move(-1)
                return True
            elif event.key == "down":
                event.prevent_default()
                event.stop()
                app._project_edit_move(1)
                return True
            elif event.key == "enter" and not self.value.strip():
                event.prevent_default()
                event.stop()
                app._project_edit_start_rename()
                return True

        # Session edit navigation
        if app and getattr(app, "_view_mode", "") == "edit_sessions" and getattr(app, "_editing_session", None) is None:
            if event.key == "up":
                event.prevent_default()
                event.stop()
                app._session_edit_index = max(0, app._session_edit_index - 1)
                app._render_edit_sessions()
                return True
            elif event.key == "down":
                event.prevent_default()
                event.stop()
                app._session_edit_index = min(len(app._viewing_sessions) - 1, app._session_edit_index + 1)
                app._render_edit_sessions()
                return True
            elif event.key == "enter" and not self.value.strip():
                event.prevent_default()
                event.stop()
                app._select_edit_sessions("")
                return True

        # Edit mode navigation: Up/Down move cursor, Enter starts rename
        if app and getattr(app, "_view_mode", "") == "edit" and getattr(app, "_editing_task", None) is None:
            if event.key == "up":
                event.prevent_default()
                event.stop()
                app._edit_move(-1)
                return True
            elif event.key == "down":
                event.prevent_default()
                event.stop()
                app._edit_move(1)
                return True
            elif event.key == "enter" and not self.value.strip():
                event.prevent_default()
                event.stop()
                app._edit_start_rename()
                return True

        if event.key == "up":
            event.prevent_default()
            event.stop()
            if self._history:
                if self._history_index == -1:
                    self._draft = self.value
                    self._history_index = len(self._history) - 1
                elif self._history_index > 0:
                    self._history_index -= 1
                self.value = self._history[self._history_index]
                self.cursor_position = len(self.value)
            return True

        elif event.key == "down":
            event.prevent_default()
            event.stop()
            if self._history_index == -1:
                pass
            elif self._history_index < len(self._history) - 1:
                self._history_index += 1
                self.value = self._history[self._history_index]
            else:
                self._history_index = -1
                self.value = self._draft
            self.cursor_position = len(self.value)
            return True

        return False


# ── Application ──────────────────────────────────────────────────────────────


class TimexApp(App):
    """Timex — time tracker."""

    TITLE = "Timex"

    CSS = """
    Screen {
        background: #171717;
        layout: vertical;
    }

    #timer {
        height: auto;
        margin: 1 2 0 2;
    }

    #history-scroll {
        height: 1fr;
        margin: 1 2 0 2;
        border: round #333333;
        border-title-color: #d4d4d4;
        border-title-style: bold;
        border-title-align: center;
        scrollbar-size: 0 0;
    }

    #history {
        height: auto;
        padding: 1 1;
    }

    #task-input {
        margin: 0 2 0 2;
        border: tall #333333;
        background: #1e1e1e;
        color: #d4d4d4;
    }

    #task-input:focus {
        border: tall #e8a55d;
    }

    #task-input > .input--placeholder {
        color: #555555;
    }


    #toast-bar {
        height: auto;
        margin: 0 2 0 2;
        color: #d4d4d4;
    }

    #footer-bar {
        height: 1;
        margin: 1 2 1 2;
        color: #555555;
        text-align: center;
    }
    """

    BINDINGS = [
        Binding("ctrl+q", "quit", "Quit", show=False),
    ]

    def _enter_view(self, mode: str, placeholder: str) -> None:
        """Switch to a sub-view (help, notification, color, dates, edit)."""
        self._view_mode = mode
        self._render_timer()
        self._render_history()
        self.query_one("#task-input", HistoryInput).placeholder = placeholder

    def _leave_view(self, toast_msg: str | None = None) -> None:
        """Return to timeline from any sub-view."""
        self._view_mode = "timeline"
        self._export_range_input = False
        if toast_msg:
            self._toast(toast_msg)
        self._mark_dirty()
        self._update_placeholder()

    @staticmethod
    def _parse_duration(raw: str) -> float:
        """Parse human duration string (e.g. '1h30m', '10min') into seconds. Returns 0 on failure."""
        parts = re.findall(
            r"(\d+)\s*(h|hr|hours?|m|min|mins?|minutes?|s|sec|secs?|seconds?)?",
            raw.lower(),
        )
        if not parts:
            return 0.0
        total = 0.0
        for value, unit in parts:
            n = int(value)
            if unit.startswith("h"):
                total += n * 3600
            elif unit.startswith("s"):
                total += n
            else:
                total += n * 60
        return total

    @staticmethod
    def _space_between(left_markup: str, right_markup: str) -> Table:
        t = Table(show_header=False, box=None, padding=0, expand=True)
        t.add_column(ratio=1)
        t.add_column(justify="right", no_wrap=True)
        t.add_row(Text.from_markup(left_markup), Text.from_markup(right_markup))
        return t

    def __init__(self) -> None:
        super().__init__()
        self.state: str = IDLE
        self.tasks: list[TaskEntry] = []
        self.session_start: datetime | None = None
        self.paused_at: datetime | None = None
        self.total_paused: timedelta = timedelta()
        self._final_active: float = 0.0
        self._project_history_secs: float = 0.0  # cached total from history
        self._project_history_loaded: bool = False
        self._last_session_tasks: list[TaskEntry] = []
        self._accent: str = DEFAULT_ACCENT
        self._accent_hex: str = DEFAULT_ACCENT_HEX
        self._last_reminder: float = 0.0
        self._last_autosave: float = 0.0
        self._dirty_history: bool = True
        self._last_saved_at: str = ""  # track saved_at to detect external changes
        self._view_mode: str = "timeline"  # "timeline" | "dates" | "history_detail" | "help" | "notification" | "color"
        self._viewing_tasks: list[TaskEntry] = []
        self._viewing_date: str = ""
        self._viewing_date_str: str = ""  # ISO date for resume
        self._dates_list: list[str] = []  # ordered date strings for selection
        self._reminder_interval: int = DEFAULT_REMINDER_INTERVAL
        self._edit_index: int = 0  # selected task index in edit mode
        self._editing_task: int | None = None  # index of task being renamed
        self._export_range_input: bool = False  # waiting for custom date range
        self._export_period: str = "today"  # today | week | month | range
        self._export_range: tuple | None = None  # (date_from, date_to) for custom range
        self._viewing_sessions: list[dict] = []
        self._session_edit_index: int = 0
        self._editing_session: int | None = None
        self._project_edit_index: int = 0  # selected project in project_edit
        self._project_editing: int | None = None  # index of project being renamed
        self._project_to_delete: str | None = None  # project name pending deletion
        self._project: str | None = None  # active project name
        self._input_wait_t: float = 0.0  # 0.0=accent, 1.0=blue (smooth transition)

        

    # ── Compose ──────────────────────────────────────────────────────────

    def compose(self) -> ComposeResult:
        yield Static(id="timer")
        scroll = VerticalScroll(Static(id="history"), id="history-scroll")
        scroll.border_title = "Timeline"
        yield scroll
        yield Static(id="toast-bar")
        yield HistoryInput(app_ref=self, placeholder="  What are you working on?  (/start to begin)", id="task-input")
        yield Static(id="footer-bar")

    # ── Project paths ──────────────────────────────────────────────────

    def _project_dir(self) -> Path:
        if self._project:
            return PROJECTS_DIR / self._project
        return STATE_DIR

    def _state_file(self) -> Path:
        return self._project_dir() / "state.json"

    def _history_file(self) -> Path:
        return self._project_dir() / "history.json"

    def _load_active_project(self) -> None:
        try:
            if ACTIVE_PROJECT_FILE.exists():
                name = ACTIVE_PROJECT_FILE.read_text().strip()
                if name:
                    self._project = name
        except OSError:
            pass

    def on_mount(self) -> None:
        self._load_active_project()
        self._load_config()
        self._load_state()
        # Apply accent color to input focus border
        if self._accent != DEFAULT_ACCENT:
            self.call_after_refresh(
                lambda: self.query_one("#task-input").styles.__setattr__("border", ("tall", self._accent))
            )
        self.set_interval(0.5, self._tick)
        self._mark_dirty()
        self._update_placeholder()
        self.call_after_refresh(lambda: self.query_one("#task-input", HistoryInput).focus())
        self._snapshot("daily", once_per_day=True)

    def on_click(self) -> None:
        self.query_one("#task-input", HistoryInput).focus()


    # ── Time helpers ─────────────────────────────────────────────────────

    def _now(self) -> datetime:
        return datetime.now()

    def _load_config(self) -> None:
        try:
            if CONFIG_FILE.exists():
                cfg = json.loads(CONFIG_FILE.read_text())
                ri = cfg.get("reminder_interval")
                if ri is not None:
                    self._reminder_interval = int(ri)
                color = cfg.get("accent_color")
                if color and re.match(r"^#[0-9a-fA-F]{6}$", color):
                    self._accent = color.lower()
                    self._accent_hex = color.lstrip("#").upper()
        # A hand-edited config must not stop the app from opening: int() on a
        # non-numeric reminder_interval raises ValueError/TypeError.
        except (OSError, json.JSONDecodeError, KeyError, ValueError, TypeError):
            pass

    @staticmethod
    def _save_config(key: str, value) -> None:
        """Update a single key in config.json (atomic write). Pass None to remove key."""
        try:
            cfg = json.loads(CONFIG_FILE.read_text()) if CONFIG_FILE.exists() else {}
        except (OSError, json.JSONDecodeError):
            cfg = {}
        if value is None:
            cfg.pop(key, None)
        else:
            cfg[key] = value
        try:
            STATE_DIR.mkdir(parents=True, exist_ok=True)
            tmp = CONFIG_FILE.with_suffix(".tmp")
            tmp.write_text(json.dumps(cfg, indent=2))
            tmp.replace(CONFIG_FILE)
        except OSError:
            pass

    # ── APM Tracking (HIDIdleTime) ──────────────────────────────────────

    # ── Time helpers ─────────────────────────────────────────────────────

    def _active_seconds(self) -> float:
        """Total active (non-paused) seconds since session start."""
        if self.state == IDLE:
            return self._final_active
        if not self.session_start:
            return 0.0

        if self.state == RUNNING:
            elapsed = (self._now() - self.session_start) - self.total_paused
        elif self.state == PAUSED and self.paused_at:
            elapsed = (self.paused_at - self.session_start) - self.total_paused
        elif self.state == PAUSED:
            self.paused_at = self._now()
            try:
                with open(CRASH_LOG, "a") as _f:
                    _f.write(f"[state] PAUSED without paused_at, auto-fixed\n")
            except OSError:
                pass
            elapsed = (self.paused_at - self.session_start) - self.total_paused
        else:
            elapsed = timedelta()

        return max(0.0, elapsed.total_seconds())

    @staticmethod
    def _fmt_time(seconds: float) -> str:
        s = int(seconds)
        h, s = divmod(s, 3600)
        m, s = divmod(s, 60)
        return f"{h:02d}:{m:02d}:{s:02d}"


    # ── External state sync ─────────────────────────────────────────────

    def _check_external_changes(self) -> None:
        """Detect if menubar (or another process) modified state.json."""
        try:
            sf = self._state_file()
            if not sf.exists():
                return
            data = json.loads(sf.read_text())
        except (OSError, json.JSONDecodeError):
            return

        file_saved_at = data.get("saved_at", "")
        if not file_saved_at or file_saved_at == self._last_saved_at:
            return

        # External change detected — reload state
        self._last_saved_at = file_saved_at
        try:
            saved_state = data.get("state", IDLE)
            self.tasks = [self._deserialize_task(d) for d in data.get("tasks", [])]
            self._last_session_tasks = [self._deserialize_task(d) for d in data.get("last_session_tasks", [])]
            self._final_active = data.get("final_active", 0.0)
            self.total_paused = timedelta(seconds=data.get("total_paused_secs", 0.0))

            session_start_str = data.get("session_start")
            self.session_start = datetime.fromisoformat(session_start_str) if session_start_str else None

            paused_at_str = data.get("paused_at")
            self.paused_at = datetime.fromisoformat(paused_at_str) if paused_at_str else None

            self.state = saved_state
            self._dirty_history = True
            if self._view_mode == "timeline":
                self._render_all()
            self._update_placeholder()
        except (KeyError, ValueError, TypeError):
            pass

    # ── Rendering ────────────────────────────────────────────────────────

    def _tick(self) -> None:
        try:
            self._check_external_changes()
            if self.state != IDLE:
                self._render_all()
                self._check_reminder()
                now = _time.monotonic()
                if now - self._last_autosave >= AUTOSAVE_INTERVAL:
                    self._last_autosave = now
                    self._save_state()
            elif self._is_input_waiting() or self._input_wait_t > 0.0:
                self._render_footer()
        except Exception:
            CRASH_LOG.parent.mkdir(parents=True, exist_ok=True)
            CRASH_LOG.write_text(traceback.format_exc())
            raise

    def _render_all(self) -> None:
        self._render_timer()
        if self._dirty_history or (self.state == RUNNING and self.tasks and self.tasks[-1].active_end is None):
            self._render_history()
            self._dirty_history = False
        self._render_footer()

    def _mark_dirty(self) -> None:
        """Mark history for re-render and trigger full render."""
        self._dirty_history = True
        self._render_all()

    def _render_timer(self) -> None:
        in_project_view = self._view_mode == "project"

        if in_project_view:
            # In /project view: show total across all projects
            active = self._all_sessions_active_seconds()
        else:
            active = self._active_seconds()
        time_str = self._fmt_time(active)

        if self.state == RUNNING:
            indicator = f"[bold {self._accent}]\u25cf[/] [bold {self._accent}]REC[/]"
            time_markup = f"[bold {self._accent}]{time_str}[/]"
        elif self.state == PAUSED:
            indicator = "[bold #888888]\u275a\u275a PAUSED[/]"
            time_markup = "[bold #888888]{0}[/]".format(time_str)
        else:
            indicator = f"[{DIM}]\u25cb  IDLE[/]"
            time_markup = f"[{DIM}]{time_str}[/]"

        status_text = Text.from_markup(f"{indicator}    {time_markup}")

        if self._project and not in_project_view:
            from rich.table import Table
            # Project total hours
            total_secs = self._project_total_seconds()
            total_h = total_secs / 3600
            if total_h >= 1:
                total_str = f"{total_h:.1f}h"
            else:
                total_str = f"{int(total_secs / 60)}m"
            # Project name + total left, status+time right
            tbl = Table(show_header=False, show_edge=False, show_lines=False,
                        padding=0, expand=True, box=None)
            tbl.add_column(ratio=1)
            tbl.add_column(justify="right")
            name = self._project
            inner_w = max(self.size.width - 12, 16)
            status_len = len(status_text)
            max_name = inner_w - status_len - len(total_str) - 4
            if len(name) > max_name:
                name = name[:max(1, max_name - 1)] + "\u2026"
            name_markup = f"[bold {TEXT_COLOR}]{name}[/] [{DIM}]{total_str}[/]"
            tbl.add_row(
                Text.from_markup(name_markup),
                status_text,
            )
            content = tbl
        else:
            content = Align.center(status_text, vertical="middle")

        panel = Panel(
            content,
            title=f"[bold {self._accent}] \u23f1  Timex [/]",
            title_align="center",
            border_style=DIMMER,
            padding=(1, 2),
        )
        self.query_one("#timer", Static).update(panel)

    def _render_history(self) -> None:
        scroll = self.query_one("#history-scroll", VerticalScroll)

        if self._view_mode == "help":
            scroll.border_title = "Help"
            self._render_help()
            return
        if self._view_mode == "notification":
            scroll.border_title = "Notifications"
            self._render_notification()
            return
        if self._view_mode == "edit":
            scroll.border_title = "Edit Tasks"
            self._render_edit()
            return
        if self._view_mode == "stats":
            scroll.border_title = "Statistics"
            self._render_stats()
            return
        if self._view_mode == "project":
            scroll.border_title = "Projects"
            self._render_project()
            return
        if self._view_mode == "project_edit":
            scroll.border_title = "Edit Projects"
            self._render_project_edit()
            return
        if self._view_mode == "confirm_delete_project":
            scroll.border_title = "Delete Project"
            self._render_confirm_delete_project()
            return
        if self._view_mode == "export":
            scroll.border_title = "Export"
            self._render_export()
            return
        if self._view_mode == "confirm_reset":
            scroll.border_title = "Reset"
            self._render_confirm_reset()
            return
        if self._view_mode == "dates":
            scroll.border_title = "History"
            self._render_dates_list()
            return
        if self._view_mode == "date_sessions":
            scroll.border_title = self._viewing_date
            self._render_date_sessions()
            return
        if self._view_mode == "edit_sessions":
            scroll.border_title = self._viewing_date
            self._render_edit_sessions()
            return
        if self._view_mode == "history_detail":
            scroll.border_title = self._viewing_date
            self._render_tasks(self._viewing_tasks, is_live=False)
            return

        scroll.border_title = "Timeline"
        display_tasks = self.tasks if self.tasks else self._last_session_tasks
        if not display_tasks:
            if self.state == RUNNING:
                self.query_one("#history", Static).update(
                    Text.from_markup(f"\n  [white]Type what you\u2019re working on[/]\n"))
            elif self.state == PAUSED:
                self.query_one("#history", Static).update(
                    Text.from_markup(f"\n  [white]Timer paused \u2014 /resume to continue[/]\n"))
            else:
                self.query_one("#history", Static).update(
                    Text.from_markup(f"\n  [white]Type a task name to start[/]\n"))
            return
        self._render_tasks(display_tasks, is_live=True)

    def _render_tasks(self, display_tasks: list[TaskEntry], is_live: bool) -> None:
        active = self._active_seconds() if is_live else None
        from rich.console import Group


        rows = []
        for i, task in enumerate(display_tasks):
            is_current = is_live and i == len(display_tasks) - 1 and task.active_end is None
            dur = task.format_duration(active if is_current else None)
            time_str = task.format_start()

            if i > 0:
                rows.append(Text.from_markup(f"[{SEPARATOR}]{'─' * 50}[/]"))

            is_thinking = task.name.startswith("\u23f3")
            if is_current:
                header = self._space_between(f"[{DIM}]{time_str}[/]", f"[bold {self._accent}]{dur} ◄[/]")
                if is_thinking:
                    name_line = Text.from_markup(f"[italic {DIM}]{task.name}[/]")
                else:
                    name_line = Text.from_markup(f"[bold {TEXT_COLOR}]{task.name}[/]")
            else:
                header = self._space_between(f"[{DIM}]{time_str}[/]", f"[#888888]{dur}[/]")
                if is_thinking:
                    name_line = Text.from_markup(f"[italic {DIM}]{task.name}[/]")
                else:
                    name_line = Text.from_markup(f"[{TEXT_COLOR}]{task.name}[/]")

            rows.append(header)
            rows.append(name_line)

        self.query_one("#history", Static).update(Group(*rows))

    def _render_dates_list(self) -> None:
        history = self._load_history()
        if not history:
            self.query_one("#history", Static).update(
                Text.from_markup(f"\n  [white]No history yet \u2014 complete a session first[/]\n")
            )
            return

        # Group by date
        by_date: dict[str, list[dict]] = {}
        for session in reversed(history):
            d = session.get("date", "unknown")
            by_date.setdefault(d, []).append(session)

        self._dates_list = list(by_date.keys())

        from rich.console import Group


        rows = []
        for i, (date_str, sessions) in enumerate(by_date.items()):
            if i > 0:
                rows.append(Text.from_markup(f"[{SEPARATOR}]{'─' * 50}[/]"))

            total = sum(s.get("total_active", 0) for s in sessions)
            task_count = sum(len(s.get("tasks", [])) for s in sessions)

            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d")
                nice_date = dt.strftime("%a, %b %d %Y")
            except ValueError:
                nice_date = date_str

            idx = i + 1
            rows.append(self._space_between(
                f"[bold {TEXT_COLOR}]{idx}. {nice_date}[/]",
                f"[{self._accent}]{self._fmt_time(total)}[/]",
            ))
            rows.append(Text.from_markup(
                f"   [{DIM}]{task_count} task{'s' if task_count != 1 else ''}[/]"
            ))

        rows.append(Text(""))
        rows.append(Text.from_markup(
            f"  [{DIM}]Enter number to view \u2022 /back to return[/]"
        ))

        self.query_one("#history", Static).update(Group(*rows))

    def _render_date_sessions(self) -> None:
        rows = []
        sessions = self._viewing_sessions
        for i, s in enumerate(sessions):
            if i > 0:
                rows.append(Text.from_markup(f"[{SEPARATOR}]{'─' * 50}[/]"))
            total = s.get("total_active", 0)
            tasks = s.get("tasks", [])
            n = len(tasks)
            label = s.get("label", f"Session {i + 1}")
            # Time range from wall_start of first task → wall_end of last task
            t_start = ""
            t_end = ""
            if tasks:
                try:
                    t_start = datetime.fromisoformat(tasks[0].get("wall_start", "")).strftime("%H:%M")
                except (ValueError, AttributeError):
                    pass
                try:
                    last = tasks[-1]
                    ws = last.get("wall_end") or last.get("wall_start", "")
                    t_end = datetime.fromisoformat(ws).strftime("%H:%M")
                except (ValueError, AttributeError):
                    pass
            time_range = f"{t_start}\u2013{t_end}" if t_start and t_end else ""
            rows.append(self._space_between(
                f"[bold {self._accent}]{i + 1}.[/] [{TEXT_COLOR}]{label}[/]",
                f"[{self._accent}]{self._fmt_time(total)}[/]",
            ))
            rows.append(Text.from_markup(
                f"   [{DIM}]{time_range}  \u2022  {n} task{'s' if n != 1 else ''}[/]"
            ))
        self.query_one("#history", Static).update(Group(*rows))

    def _select_date_sessions(self, raw: str) -> None:
        if not raw.isdigit():
            return
        num = int(raw)
        if num < 1 or num > len(self._viewing_sessions):
            self._toast(f"Enter 1\u2013{len(self._viewing_sessions)}")
            return
        session = self._viewing_sessions[num - 1]
        tasks = [self._deserialize_task(td) for td in session.get("tasks", [])]
        if not tasks:
            self._toast("No tasks in this session")
            return
        self._view_mode = "history_detail"
        self._viewing_tasks = tasks
        self._render_history()
        inp = self.query_one("#task-input", HistoryInput)
        inp.placeholder = "  /edit to manage \u2022 /back to sessions"

    # ── Session edit (rename/delete from history) ────────────────────────

    def _cmd_edit_sessions(self) -> None:
        if not self._viewing_sessions:
            self._toast("No sessions to edit")
            return
        self._view_mode = "edit_sessions"
        self._session_edit_index = 0
        self._editing_session = None
        self._render_history()
        inp = self.query_one("#task-input", HistoryInput)
        inp.placeholder = "  \u2191/\u2193 to select \u2022 Enter to rename \u2022 /back"

    def _render_edit_sessions(self) -> None:
        rows = []
        sessions = self._viewing_sessions
        for i, s in enumerate(sessions):
            if i > 0:
                rows.append(Text.from_markup(f"[{SEPARATOR}]{'─' * 50}[/]"))
            total = s.get("total_active", 0)
            n = len(s.get("tasks", []))
            label = s.get("label", f"Session {i + 1}")
            selected = (i == self._session_edit_index)
            if selected:
                rows.append(Text.from_markup(
                    f"[bold {self._accent}]\u25b6 {label}[/]  [{DIM}]{self._fmt_time(total)} \u2022 {n} tasks[/]"
                ))
            else:
                rows.append(Text.from_markup(
                    f"  [{TEXT_COLOR}]{label}[/]  [{DIM}]{self._fmt_time(total)} \u2022 {n} tasks[/]"
                ))
        rows.append(Text(""))
        rows.append(Text.from_markup(f"  [{DIM}]Enter to rename \u2022 empty to delete \u2022 /back[/]"))
        self.query_one("#history", Static).update(Group(*rows))

    def _select_edit_sessions(self, raw: str) -> None:
        if self._editing_session is not None:
            # Submitting rename or delete
            idx = self._editing_session
            if idx >= len(self._viewing_sessions):
                self._editing_session = None
                return
            if raw:
                # Rename: set label
                self._viewing_sessions[idx]["label"] = raw
                self._save_sessions_to_history()
                self._toast("Session renamed")
            else:
                # Delete session
                if not self._snapshot("session-delete"):
                    self._toast("Backup failed — session kept", 5)
                    self._editing_session = None
                    return
                self._viewing_sessions.pop(idx)
                self._save_sessions_to_history()
                self._toast("Session deleted")
                if not self._viewing_sessions:
                    self._leave_view("All sessions deleted")
                    return
                self._session_edit_index = min(self._session_edit_index, len(self._viewing_sessions) - 1)
            self._editing_session = None
            self._render_edit_sessions()
            inp = self.query_one("#task-input", HistoryInput)
            inp.placeholder = "  \u2191/\u2193 to select \u2022 Enter to rename \u2022 /back"
            return

        # Enter on selected → start rename
        idx = self._session_edit_index
        self._editing_session = idx
        label = self._viewing_sessions[idx].get("label", f"Session {idx + 1}")
        inp = self.query_one("#task-input", HistoryInput)
        inp.value = label
        inp.placeholder = "  New name (empty to delete) \u2022 Enter to confirm"

    def _save_sessions_to_history(self) -> None:
        """Rewrite history.json with updated/deleted sessions for the viewed date."""
        history = self._load_history()
        date_str = self._viewing_date_str
        # Remove all sessions for this date
        history = [s for s in history if s.get("date") != date_str]
        # Re-add the remaining (possibly modified) sessions
        history.extend(self._viewing_sessions)
        # Sort by date + session_start
        history.sort(key=lambda s: (s.get("date", ""), s.get("session_start", "")))
        hf = self._history_file()
        hf.parent.mkdir(parents=True, exist_ok=True)
        tmp = hf.with_suffix(".tmp")
        tmp.write_text(json.dumps(history, indent=2))
        tmp.replace(hf)
        self._invalidate_history_cache()

    # Smooth gradient keyframes for update border

    def _is_input_waiting(self) -> bool:
        """True when app is waiting for freeform text from user."""
        if self._export_range_input:
            return True
        if self._view_mode == "edit" and self._editing_task is not None:
            return True
        if self._view_mode == "project_edit" and self._project_editing is not None:
            return True
        if self._view_mode == "notification":
            return True
        return False

    def _waiting_border_color(self) -> str:
        """Interpolate accent → blue based on _input_wait_t (0.0–1.0)."""
        # Parse accent hex
        a = self._accent.lstrip("#")
        r1, g1, b1 = int(a[0:2], 16), int(a[2:4], 16), int(a[4:6], 16)
        r2, g2, b2 = 0x61, 0xaf, 0xef  # #61afef Blue
        t = self._input_wait_t
        r = int(r1 + (r2 - r1) * t)
        g = int(g1 + (g2 - g1) * t)
        b = int(b1 + (b2 - b1) * t)
        return f"#{r:02x}{g:02x}{b:02x}"

    def _render_footer(self) -> None:
        # Animate input border: accent ↔ blue based on waiting state
        waiting = self._is_input_waiting()
        step = 0.15  # per tick (0.5s) → ~3s full transition
        if waiting:
            self._input_wait_t = min(1.0, self._input_wait_t + step)
        else:
            self._input_wait_t = max(0.0, self._input_wait_t - step)

        inp = self.query_one("#task-input", HistoryInput)
        if self._input_wait_t > 0.0:
            inp.styles.border = ("tall", self._waiting_border_color())
        else:
            inp.styles.border = ("tall", self._accent)
        today = self._now().strftime("%a, %b %d %Y")
        parts = [f"[{DIM}]{today}[/]"]
        footer = Text.from_markup("  ".join(parts))
        footer.justify = "center"
        self.query_one("#footer-bar", Static).update(footer)

    # ── Input handling ───────────────────────────────────────────────────

    @on(Input.Submitted)
    def _on_submit(self, event: Input.Submitted) -> None:
      try:
        raw = event.value.strip()
        inp = self.query_one("#task-input", HistoryInput)
        inp.value = ""

        # The widget writes state.json too, and the tick only notices up to half a
        # second later. Acting on stale memory and saving it would silently undo
        # whatever was just clicked in the menu bar, so pull its changes in first.
        self._check_external_changes()

        if not raw:
            if self._view_mode == "edit" and self._editing_task is not None:
                self._submit_edit("")
            elif self._view_mode == "project_edit" and self._project_editing is not None:
                self._select_project_edit("")
            return

        # Transliterate cyrillic → latin for commands
        if raw.startswith("/"):
            raw = _translit_cmd(raw)

        # Resolve partial commands: /b → /back, /pa → /pause, etc.
        if raw.startswith("/") and " " not in raw:
            val = raw.lower()
            if self._view_mode != "timeline":
                if "/back".startswith(val) and val != "/back":
                    raw = "/back"
                elif self._view_mode == "project" and "/edit".startswith(val) and val != "/edit":
                    raw = "/edit"
            else:
                commands = STATE_COMMANDS.get(self.state, STATE_COMMANDS[IDLE])
                for c in commands:
                    if c.startswith(val) and val != c:
                        raw = c
                        break

        inp.add_to_history(raw)

        cmd = raw.lower()

        if cmd == "/start":
            self._cmd_start()
        elif cmd == "/pause":
            self._cmd_pause()
        elif cmd == "/resume":
            self._cmd_resume()
        elif cmd == "/export":
            self._cmd_export()
        elif cmd == "/new":
            self._cmd_new()
        elif cmd == "/clear":
            self._cmd_clear()
        elif cmd.startswith("/add"):
            self._cmd_add_time(raw)
        elif cmd.startswith("/remove"):
            self._cmd_remove_time(raw)
        elif cmd == "/date":
            self._cmd_date()
        elif cmd == "/help":
            self._cmd_help()
        elif cmd == "/notification":
            self._cmd_notification()
        elif cmd == "/edit":
            self._cmd_edit()
        elif cmd == "/stats":
            self._cmd_stats()
        elif cmd == "/back":
            self._cmd_back()
        elif cmd == "/reset":
            self._cmd_reset()
        elif cmd == "/project":
            self._cmd_project()
        elif self._view_mode == "edit":
            self._submit_edit(raw)
        elif self._view_mode == "dates" and raw.isdigit():
            self._select_date(int(raw))
        elif self._view_mode == "date_sessions":
            self._select_date_sessions(raw)
        elif self._view_mode == "edit_sessions":
            self._select_edit_sessions(raw)
        elif self._view_mode == "notification":
            self._select_notification(raw)
        elif self._view_mode == "project":
            self._select_project(raw)
        elif self._view_mode == "project_edit":
            self._select_project_edit(raw)
        elif self._view_mode == "confirm_delete_project":
            self._select_confirm_delete_project(raw)
        elif self._view_mode == "confirm_reset":
            self._select_confirm_reset(raw)
        elif self._view_mode == "export":
            self._select_export(raw)
        elif raw.startswith("/"):
            self._toast(f"Unknown command: {raw}")
        else:
            self._add_task(raw)
      except Exception:
        CRASH_LOG.parent.mkdir(parents=True, exist_ok=True)
        CRASH_LOG.write_text(traceback.format_exc())
        raise

    # ── Commands ─────────────────────────────────────────────────────────

    def _cmd_start(self) -> None:
        if self.state == PAUSED:
            self._cmd_resume()
            return
        if self.state == RUNNING:
            # Already running — build history entry BEFORE resetting state
            entry = self._build_history_entry()
            self.state = IDLE
            self.tasks = []
            self._last_session_tasks = []
            self.session_start = None
            self.paused_at = None
            self.total_paused = timedelta()
            self._final_active = 0.0
            self._view_mode = "timeline"
            self._save_state()  # persist clean state first
            if entry:
                self._append_history(entry)  # then append history
            self._toast("Session saved \u2014 new day started")
            self._update_placeholder()
            self._mark_dirty()
            return

        # IDLE — start the timer
        self.state = RUNNING
        self.session_start = self._now()
        self.total_paused = timedelta()
        self.paused_at = None
        self._final_active = 0.0
        self.tasks = []
        self._view_mode = "timeline"
        self._reset_reminder()
        self._toast("Timer started")
        self._update_placeholder()
        self._mark_dirty()
        self._save_state()

    def _cmd_pause(self) -> None:
        if self.state != RUNNING:
            self._toast("Timer is not running")
            return

        self.state = PAUSED
        self.paused_at = self._now()
        self._reset_reminder()
        self._toast("Timer paused")
        self._update_placeholder()
        self._mark_dirty()
        self._save_state()


    def _cmd_resume(self) -> None:
        # Can't resume from history view — use /start instead
        if self._view_mode == "history_detail" and self.state == IDLE:
            self._toast("Use /start to begin a new session")
            return

        if self.state != PAUSED:
            self._toast("Timer is not paused")
            return

        if self.paused_at is None:
            self.paused_at = self._now()
        pause_dur = self._now() - self.paused_at
        self.total_paused += pause_dur
        self.paused_at = None
        self.state = RUNNING

        self._reset_reminder()
        self._toast("Timer resumed")
        self._update_placeholder()
        self._mark_dirty()
        self._save_state()

    def _cmd_reset(self) -> None:
        """Reset current session — ask for confirmation."""
        if self.state == IDLE:
            self._toast("Nothing to reset")
            return
        self._enter_view("confirm_reset", "  y to confirm, n to cancel")

    def _render_confirm_reset(self) -> None:
        rows = [
            Text(""),
            Text.from_markup(f"[bold {self._accent}]Reset session?[/]"),
            Text(""),
            Text.from_markup(f"[{DIM}]This will stop the timer and clear all tasks.[/]"),
            Text.from_markup(f"[{DIM}]The session will NOT be saved to history.[/]"),
            Text(""),
            Text.from_markup(f"[{TEXT_COLOR}]y[/][{DIM}] — confirm reset[/]"),
            Text.from_markup(f"[{TEXT_COLOR}]n[/][{DIM}] — cancel[/]"),
        ]
        self.query_one("#history", Static).update(Group(*rows))

    def _select_confirm_reset(self, raw: str) -> None:
        if raw.lower() not in ("y", "yes", "n", "no"):
            return
        if raw.lower() in ("y", "yes"):
            self.state = IDLE
            self.tasks = []
            self._last_session_tasks = []
            self.session_start = None
            self.paused_at = None
            self.total_paused = timedelta()
            self._final_active = 0.0
            self._leave_view("Session reset")
            self._save_state()
        else:
            self._leave_view("Reset cancelled")

    # ── Watch (window activity monitor) ───────────────────────────────────


    # ── Prompt/dialog detection ──────────────────────────────────────────

    def _add_task(self, name: str) -> None:
        if self.state == IDLE:
            # Auto-start timer
            self.state = RUNNING
            self.session_start = self._now()
            self.total_paused = timedelta()
            self.paused_at = None
            self._final_active = 0.0
            self.tasks = []
            self._view_mode = "timeline"
            self._reset_reminder()
        elif self.state == PAUSED:
            # Auto-resume on new task
            self.total_paused += self._now() - self.paused_at
            self.paused_at = None
            self.state = RUNNING
        elif self.state != RUNNING:
            return

        now = self._now()
        active = self._active_seconds()

        # Finalize previous task
        if self.tasks and self.tasks[-1].active_end is None:
            self.tasks[-1].active_end = active
            self.tasks[-1].wall_end = now

        self.tasks.append(TaskEntry(
            name=name,
            wall_start=now,
            active_start=active,
        ))

        self._mark_dirty()
        self._save_state()

        # Scroll to bottom
        scroll = self.query_one("#history-scroll", VerticalScroll)
        scroll.scroll_end(animate=False)

    # ── /export — Export ─────────────────────────────────────────────────────

    _EXPORT_PERIODS = (
        ("today", "Today"),
        ("week", "This week"),
        ("month", "This month"),
        ("range", "Custom range"),
    )

    def _cmd_export(self) -> None:
        """Open export view: pick a period, save an .xlsx report."""
        if self._view_mode == "history_detail" and self._viewing_date_str:
            try:
                d = datetime.strptime(self._viewing_date_str, "%Y-%m-%d").date()
                self._export_period = "range"
                self._export_range = (d, d)
            except ValueError:
                self._export_period = "today"
        if self._export_period == "range" and not self._export_range:
            self._export_period = "today"
        self._export_range_input = False
        self._enter_view("export", "  select option • /back")

    def _period_bounds(self, period: str) -> tuple[date, date]:
        """Inclusive (from, to) for a period key."""
        today = self._now().date()
        if period == "week":
            return today - timedelta(days=today.weekday()), today
        if period == "month":
            return today.replace(day=1), today
        if period == "range" and self._export_range:
            return self._export_range
        return today, today

    def _period_label(self, period: str) -> str:
        d1, d2 = self._period_bounds(period)
        if d1 == d2:
            return d1.strftime("%A, %B ") + str(d1.day) + d1.strftime(", %Y")
        if (d1.year, d1.month) == (d2.year, d2.month):
            return d1.strftime("%B ") + f"{d1.day}–{d2.day}" + d2.strftime(", %Y")
        return (d1.strftime("%b ") + str(d1.day) + " – "
                + d2.strftime("%b ") + str(d2.day) + d2.strftime(", %Y"))

    def _collect_tasks(self, date_from: date, date_to: date) -> list[TaskEntry]:
        """Tasks starting within [date_from, date_to], chronological.

        History only gains a session once it is closed with /new, so a stopped
        session still lives in memory and has to be picked up from there.
        """
        out: list[TaskEntry] = []
        for session in self._load_history():
            for td in session.get("tasks", []):
                t = self._deserialize_task(td)
                if date_from <= t.wall_start.date() <= date_to:
                    out.append(t)
        live = self.tasks if self.tasks else self._last_session_tasks
        for t in live:
            if date_from <= t.wall_start.date() <= date_to:
                out.append(t)
        out.sort(key=lambda t: t.wall_start)
        return out

    def _task_secs(self, t: TaskEntry, active: float) -> float:
        return t.get_duration(active if t.active_end is None else None)


    def _period_tasks(self, period: str) -> tuple[list[TaskEntry], float]:
        tasks = self._collect_tasks(*self._period_bounds(period))
        active = self._active_seconds()
        return tasks, sum(self._task_secs(t, active) for t in tasks)

    def _render_export(self) -> None:
        """Export view: period picker plus a summary of what would be exported."""
        if self._export_range_input:
            self._render_export_range()
            return

        tasks, total = self._period_tasks(self._export_period)
        days = len({t.wall_start.date() for t in tasks})

        rows = [Text.from_markup(
            f"[bold {self._accent}]Report on Hours[/] [{DIM}]/ {self._report_author()} for[/]"
            f" [{TEXT_COLOR}]{self._project}[/]"
        )]
        rows.append(Text.from_markup(f"[{SEPARATOR}]{'─' * 50}[/]"))
        rows.append(Text.from_markup(
            f"[{DIM}]Period:[/]  [{TEXT_COLOR}]{self._period_label(self._export_period)}[/]"
        ))
        rows.append(Text.from_markup(f"[{DIM}]Days:[/]    [{TEXT_COLOR}]{days}[/]"))
        rows.append(Text.from_markup(f"[{DIM}]Tasks:[/]   [{TEXT_COLOR}]{len(tasks)}[/]"))
        rows.append(Text.from_markup(
            f"[{DIM}]Total:[/]   [{TEXT_COLOR}]{self._fmt_time(total)}[/]"
            f"  [{DIM}]({total / 3600:.2f} h)[/]"
        ))

        for i, (key, label) in enumerate(self._EXPORT_PERIODS, 1):
            rows.append(Text.from_markup(f"[{SEPARATOR}]{'─' * 50}[/]"))
            marker = f" [{self._accent}]•[/]" if key == self._export_period else ""
            left = f"[bold {self._accent}]{i}.[/] [{TEXT_COLOR}]{label}[/]{marker}"
            if key == "range":
                rows.append(Text.from_markup(left))
            else:
                _, secs = self._period_tasks(key)
                right = f"[{TEXT_COLOR}]{self._fmt_time(secs)}[/]" if secs else f"[{DIM}]—[/]"
                rows.append(self._space_between(left, right))

        rows.append(Text.from_markup(f"[{SEPARATOR}]{'─' * 50}[/]"))
        rows.append(Text.from_markup(
            f"[bold {self._accent}]5.[/] [{TEXT_COLOR}]Share a link[/]"
        ))
        rows.append(Text.from_markup(f"[{SEPARATOR}]{'─' * 50}[/]"))
        rows.append(Text.from_markup(
            f"[bold {self._accent}]6.[/] [{TEXT_COLOR}]Export to Excel (.xlsx)[/]"
        ))
        self.query_one("#history", Static).update(Group(*rows))

    def _render_export_range(self) -> None:
        rows = [
            Text(""),
            Text.from_markup(f"[bold {self._accent}]Custom range[/]"),
            Text(""),
            Text.from_markup(
                f"[{DIM}]Two dates:[/] [{TEXT_COLOR}]2026-07-01 2026-07-15[/]"
            ),
            Text.from_markup(f"[{DIM}]One date exports that single day.[/]"),
            Text(""),
            Text.from_markup(f"[{DIM}]/back to cancel[/]"),
        ]
        self.query_one("#history", Static).update(Group(*rows))

    def _select_export(self, raw: str) -> None:
        if self._export_range_input:
            self._apply_export_range(raw)
            return
        if raw in ("1", "2", "3", "4"):
            key = self._EXPORT_PERIODS[int(raw) - 1][0]
            if key == "range":
                self._export_range_input = True
                self._enter_view("export", "  e.g. 2026-07-01 2026-07-15")
                return
            self._export_period = key
            self._render_export()
        elif raw == "5":
            self._publish_report()
        elif raw == "6":
            self._export_xlsx()

    def _apply_export_range(self, raw: str) -> None:
        parsed: list[date] = []
        for part in raw.replace(",", " ").split():
            for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m"):
                try:
                    d = datetime.strptime(part, fmt).date()
                except ValueError:
                    continue
                if fmt == "%d.%m":
                    d = d.replace(year=self._now().year)
                parsed.append(d)
                break
        if not parsed or len(parsed) > 2:
            self._toast("Use: 2026-07-01 2026-07-15", 4)
            return
        self._export_range = (min(parsed), max(parsed))
        self._export_period = "range"
        self._export_range_input = False
        self._enter_view("export", "  select option • /back")

    def _export_gather(self) -> tuple | None:
        """(from, to, tasks, durations, total) for the chosen period, or None."""
        d_from, d_to = self._period_bounds(self._export_period)
        tasks = self._collect_tasks(d_from, d_to)
        if not tasks:
            self._toast("Nothing to export for this period")
            return None
        active = self._active_seconds()
        durations = [self._task_secs(t, active) for t in tasks]
        return d_from, d_to, tasks, durations, sum(durations)

    def _report_author(self) -> str:
        """Name the report is signed with.

        Set `report_name` in config.json; otherwise fall back to the macOS
        account's full name, so someone else's export is not signed with mine.
        """
        try:
            cfg = json.loads(CONFIG_FILE.read_text()) if CONFIG_FILE.exists() else {}
            name = str(cfg.get("report_name") or "").strip()
            if name:
                return name
        except (OSError, json.JSONDecodeError, TypeError, ValueError):
            pass
        try:
            import pwd
            full = pwd.getpwuid(os.getuid()).pw_gecos.split(",")[0].strip()
            if full:
                return full
        except Exception:
            pass
        return "Timex"

    def _export_filename(self, d_from: date, d_to: date, ext: str) -> str:
        stamp = (d_from.strftime("%Y%m%d") if d_from == d_to
                 else f"{d_from.strftime('%Y%m%d')}-{d_to.strftime('%Y%m%d')}")
        slug = re.sub(r"[^a-z0-9]+", "_", (self._project or "timex").lower()).strip("_")
        parts = self._report_author().split()
        who = re.sub(r"[^a-z0-9]+", "_", (parts[-1] if parts else "timex").lower()).strip("_")
        return f"{who or 'timex'}_{slug}_{stamp}.{ext}"

    def _top_tasks(self, tasks, durations, total, limit: int) -> list[tuple]:
        """(name, seconds, share) for the biggest tasks, largest first.

        Task names are near-unique, so there is nothing to fold a tail into: an
        "Other" bucket would swallow ~80% of a month and say nothing. These are
        the biggest single tasks, not a breakdown of the whole.
        """
        agg: dict[str, float] = {}
        for t, secs in zip(tasks, durations):
            agg[t.name] = agg.get(t.name, 0.0) + secs
        ranked = sorted(agg.items(), key=lambda kv: kv[1], reverse=True)[:limit]
        return [(n, s, (s / total if total else 0.0)) for n, s in ranked]

    def _export_by_day(self, tasks, durations, d_from: date, d_to: date) -> list[tuple]:
        """(date, seconds) for every calendar day in the period, gaps included."""
        acc: dict[date, float] = {}
        for t, secs in zip(tasks, durations):
            d = t.wall_start.date()
            acc[d] = acc.get(d, 0.0) + secs
        out, d = [], d_from
        while d <= d_to:
            out.append((d, acc.get(d, 0.0)))
            d += timedelta(days=1)
        return out

    def _export_xlsx(self) -> None:
        """Write the .xlsx report for the selected period into ~/Downloads."""
        got = self._export_gather()
        if not got:
            return
        try:
            wb = self._build_workbook(*got)
        except ImportError:
            self._toast("Export unavailable: openpyxl is missing", 5)
            return
        path = Path.home() / "Downloads" / self._export_filename(got[0], got[1], "xlsx")
        path.parent.mkdir(exist_ok=True)
        wb.save(str(path))
        self._leave_view(f"Saved → ~/Downloads/{path.name}")

    # Light is the base and dark is the override, so a client who never touched
    # their OS settings gets the readable one. Both palettes are validated against
    _REPORT_CSS = """
:root{
 --surface:#fcfcfb;--panel:#fff;--line:#e7e7e3;--hair:#f0f0ec;
 --ink:#14140f;--ink2:#5c5c55;
 --s1:#b87400;--onS1:#fff;--ctx:#e7e7e3;--tipbg:#14140f;--tipink:#fcfcfb;
 color-scheme:light}
@media (prefers-color-scheme:dark){:root{
 --surface:#171717;--panel:#1e1e1e;--line:#2c2c2c;--hair:#242424;
 --ink:#e8e8e6;--ink2:#a3a39d;
 --s1:#c98500;--onS1:#171717;--ctx:#2c2c2c;--tipbg:#000;--tipink:#e8e8e6;
 color-scheme:dark}}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--surface);color:var(--ink);-webkit-font-smoothing:antialiased;
 font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Inter,Roboto,"Helvetica Neue",Arial,sans-serif;
 font-size:16px;line-height:1.6;padding:48px 24px 96px}
.num{font-variant-numeric:tabular-nums;font-feature-settings:"tnum" 1}
.wrap{max-width:1000px;margin:0 auto}
.brand{color:var(--ink2);font-size:14px;font-weight:600;letter-spacing:.04em}
h1{font-size:34px;font-weight:700;margin:2px 0 1px;letter-spacing:-.02em;line-height:1.15}
.sub{color:var(--ink2);font-size:16px;line-height:1.4}
.rule{height:1px;background:var(--line);margin:36px 0}
h2{font-size:20px;font-weight:650;margin-bottom:0;letter-spacing:-.01em;line-height:1.3}
.note{color:var(--ink2);font-size:14px;line-height:1.45;margin-bottom:26px}
.tiles{display:grid;grid-template-columns:repeat(3,1fr);gap:12px}
@media (max-width:720px){.tiles{grid-template-columns:repeat(2,1fr)}}
@media (max-width:440px){.tiles{grid-template-columns:1fr}}
.tile{background:var(--panel);border:1px solid var(--line);border-radius:12px;padding:20px 18px}
.tile .k{color:var(--ink2);font-size:14px;line-height:1.35}
.tile .v{font-size:26px;font-weight:650;margin-top:6px;white-space:nowrap;
 letter-spacing:-.02em;line-height:1.2}
.tile .n{color:var(--ink2);font-size:14px;margin-top:5px;line-height:1.35}
.tile.hero .v{font-size:34px;color:var(--s1)}
.tt{list-style:none}
.tt li{padding:11px 0;border-bottom:1px solid var(--hair);opacity:0;transform:translateY(4px);
 transition:opacity .4s ease,transform .4s ease}
.tt li.in{opacity:1;transform:none}
.tth{display:flex;gap:16px;align-items:baseline;margin-bottom:7px}
.ttn{flex:1;min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.ttv{color:var(--ink2);font-size:15px;white-space:nowrap}
.ttv b{color:var(--ink);font-weight:600;margin-left:10px}
.ttb{height:10px;background:var(--hair);border-radius:5px;overflow:hidden}
.ttb i{display:block;width:0;height:100%;background:var(--s1);border-radius:5px;
 transition:width .8s cubic-bezier(.22,1,.36,1)}
.chart{position:relative;padding-top:22px}
.peak{position:absolute;top:0;left:0;right:0;border-top:1px dashed var(--line);
 color:var(--ink2);font-size:13px;padding-bottom:3px}
.bars{display:flex;align-items:flex-end;gap:5px;height:184px;border-bottom:1px solid var(--line)}
.bar{flex:1;min-width:0;display:flex;flex-direction:column;justify-content:flex-end;height:100%}
.bar i{display:block;height:0;background:var(--s1);border-radius:4px 4px 0 0;
 transition:height .7s cubic-bezier(.22,1,.36,1)}
.bar.ctx i{background:var(--ctx)}
.bar.zero i{background:var(--hair)}
.bar:hover i{filter:brightness(1.15)}
.xaxis{display:flex;gap:5px;margin-top:10px;color:var(--ink2);font-size:13px}
.xaxis span{flex:1;min-width:0;text-align:center;white-space:nowrap;overflow:hidden}
.xaxis span.on{color:var(--ink);font-weight:600}
table{width:100%;border-collapse:collapse;font-size:16px}
th,td{padding-right:18px;white-space:nowrap}
th:last-child,td:last-child{padding-right:0}
th{color:var(--ink2);font-size:14px;font-weight:600;text-align:left;
 padding-bottom:10px;border-bottom:1px solid var(--line)}
td{padding-top:11px;padding-bottom:11px;border-bottom:1px solid var(--hair);vertical-align:top}
td.n,th.n{text-align:right}
td.tk,th.tk{white-space:normal;width:100%;min-width:240px}
.dl{position:fixed;right:28px;bottom:28px;width:58px;height:58px;border-radius:18px;
 display:flex;align-items:center;justify-content:center;background:var(--s1);
 color:var(--onS1);text-decoration:none;z-index:8;
 box-shadow:0 8px 28px rgba(0,0,0,.28);
 transition:transform .15s ease,filter .15s ease}
.dl:hover{filter:brightness(1.1);transform:translateY(-2px)}
.dl:active{transform:translateY(0)}
.dl svg{width:24px;height:24px}
footer{color:var(--ink2);font-size:14px;margin-top:36px}
#tip{position:fixed;pointer-events:none;opacity:0;background:var(--tipbg);
 border-radius:8px;padding:8px 11px;font-size:14px;color:var(--tipink);
 transition:opacity .12s;z-index:9;white-space:nowrap}
@media (prefers-reduced-motion:reduce){*{transition:none!important}}
"""

    _REPORT_JS = """
var tip=document.getElementById('tip');
document.querySelectorAll('[data-tip]').forEach(function(el){
  el.addEventListener('mouseenter',function(){tip.textContent=el.dataset.tip;tip.style.opacity=1});
  el.addEventListener('mousemove',function(e){
    var x=e.clientX+14,y=e.clientY+14;
    if(x+tip.offsetWidth>innerWidth-8){x=e.clientX-tip.offsetWidth-14}
    tip.style.left=x+'px';tip.style.top=y+'px'});
  el.addEventListener('mouseleave',function(){tip.style.opacity=0});
});
function paint(){
  document.querySelectorAll('.tt li').forEach(function(el,i){
    setTimeout(function(){el.classList.add('in');
      var b=el.querySelector('.ttb i');if(b)b.style.width=b.dataset.w},90+i*70)});
  document.querySelectorAll('[data-h]').forEach(function(el,i){
    setTimeout(function(){el.style.height=el.dataset.h},220+i*16)});
  var hero=document.getElementById('hero');
  if(!hero)return;
  var end=parseFloat(hero.dataset.secs),t0=null,done=false;
  function pad(n){return String(n).padStart(2,'0')}
  // Truncate, matching _fmt_time — rounding here would drift the hero a second
  // away from the legend, the table and the workbook.
  function f(s){s=Math.floor(s);return pad(Math.floor(s/3600))+':'+pad(Math.floor(s%3600/60))+':'+pad(s%60)}
  function land(){done=true;hero.textContent=f(end)}
  function step(ts){if(done)return;if(!t0)t0=ts;var p=Math.min(1,(ts-t0)/800);
    if(p>=1){land();return}
    hero.textContent=f(end*(1-Math.pow(1-p,3)));requestAnimationFrame(step)}
  requestAnimationFrame(step);
  // The count-up must never own the truth. If rAF timestamps stall (background tab,
  // throttling) the loop would keep repainting a partial billable total forever, so
  // this both lands the real value and stops the loop from overwriting it.
  setTimeout(land,900);
}
if(matchMedia('(prefers-reduced-motion:reduce)').matches){
  document.querySelectorAll('[data-h]').forEach(function(el){el.style.height=el.dataset.h});
  document.querySelectorAll('.ttb i').forEach(function(el){el.style.width=el.dataset.w});
  document.querySelectorAll('.tt li').forEach(function(el){el.classList.add('in')});
}else{addEventListener('load',paint)}
"""

    def _build_workbook(self, d_from, d_to, tasks, durations, total):
        """Two sheets: Report (summary + pie + bar) and Detail (one row per task)."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, Reference

        header_font = Font(bold=True, color="FFFFFF", size=11)
        accent_fill = PatternFill(start_color=self._accent_hex,
                                  end_color=self._accent_hex, fill_type="solid")
        bold = Font(bold=True, size=11)
        grey = Font(color="888888", size=10)
        thin = Border(bottom=Side(style="thin", color="DDDDDD"))
        centered = Alignment(horizontal="center")

        def _head(ws, row: int, titles: list) -> None:
            for col, title in enumerate(titles, 1):
                c = ws.cell(row=row, column=col, value=title)
                c.font = header_font
                c.fill = accent_fill
                c.alignment = centered

        def _text(ws, row: int, col: int, value: str):
            """Write a task/project name as text.

            openpyxl types a leading '=' as a formula, so a task called "=1+1" —
            or worse, a DDE payload — would execute in the client's Excel.
            """
            c = ws.cell(row=row, column=col, value=value)
            if c.data_type == "f":
                c.data_type = "s"
            return c

        wb = Workbook()

        # ── Report sheet: summary + breakdown + charts ──
        ws = wb.active
        ws.title = "Report"
        ws["A1"] = "⏱ Time Report"
        ws["A1"].font = Font(bold=True, size=14)
        _text(ws, 2, 1, f"{self._project or 'Timex'} · {self._period_label(self._export_period)}").font = grey

        by_day: dict[date, float] = {}
        for t, secs in zip(tasks, durations):
            d = t.wall_start.date()
            by_day[d] = by_day.get(d, 0.0) + secs
        busiest = max(by_day.items(), key=lambda kv: kv[1])
        avg = total / len(by_day)
        longest = max(durations)

        ws["A4"] = "Summary"
        ws["A4"].font = bold
        summary = [
            ("Total", self._fmt_time(total), total / 3600),
            ("Days worked", len(by_day), None),
            ("Tasks", len(tasks), None),
            ("Avg / day", self._fmt_time(avg), avg / 3600),
            ("Longest task", self._fmt_time(longest), longest / 3600),
            ("Busiest day", busiest[0].strftime("%a, %b ") + str(busiest[0].day),
             busiest[1] / 3600),
        ]
        for i, (label, value, hours) in enumerate(summary, start=5):
            ws.cell(row=i, column=1, value=label).font = grey
            ws.cell(row=i, column=2, value=value)
            if hours is not None:
                c = ws.cell(row=i, column=3, value=round(hours, 2))
                c.number_format = "0.00"
                c.font = grey

        # Biggest tasks, not a breakdown: names are near-unique, so a pie would put
        # ~80% of a month into one "Other" wedge and say nothing.
        ranked = self._top_tasks(tasks, durations, total, 10)

        top = 12
        ws.cell(row=top, column=1, value="Longest tasks").font = bold
        _head(ws, top + 1, ["Task", "Duration", "Hours", "Share"])
        for i, (name, secs, share) in enumerate(ranked):
            r = top + 2 + i
            _text(ws, r, 1, name)
            ws.cell(row=r, column=2, value=self._fmt_time(secs))
            c = ws.cell(row=r, column=3, value=round(secs / 3600, 2))
            c.number_format = "0.00"
            c = ws.cell(row=r, column=4, value=share)
            c.number_format = "0.0%"
            for col in range(1, 5):
                ws.cell(row=r, column=col).border = thin
        task_last = top + 1 + len(ranked)

        task_bar = BarChart()
        task_bar.type = "bar"  # horizontal: long task names need the room
        task_bar.title = "Longest tasks"
        task_bar.add_data(Reference(ws, min_col=3, min_row=top + 1, max_row=task_last),
                          titles_from_data=True)
        task_bar.set_categories(Reference(ws, min_col=1, min_row=top + 2, max_row=task_last))
        task_bar.height, task_bar.width = 9.5, 16
        task_bar.legend = None
        ws.add_chart(task_bar, "F4")

        day_top = task_last + 2
        ws.cell(row=day_top, column=1, value="By day").font = bold
        _head(ws, day_top + 1, ["Date", "Duration", "Hours"])
        r = day_top + 2
        d = d_from
        while d <= d_to:
            secs = by_day.get(d, 0.0)
            ws.cell(row=r, column=1, value=d.strftime("%a %d %b"))
            ws.cell(row=r, column=2, value=self._fmt_time(secs))
            c = ws.cell(row=r, column=3, value=round(secs / 3600, 2))
            c.number_format = "0.00"
            for col in range(1, 4):
                ws.cell(row=r, column=col).border = thin
            d += timedelta(days=1)
            r += 1
        day_last = r - 1

        if day_last > day_top + 2:  # a bar chart of one day is pointless
            bar = BarChart()
            bar.type = "col"
            bar.title = "Hours per day"
            bar.add_data(Reference(ws, min_col=3, min_row=day_top + 1, max_row=day_last),
                         titles_from_data=True)
            bar.set_categories(Reference(ws, min_col=1, min_row=day_top + 2, max_row=day_last))
            bar.height, bar.width = 8, 18
            bar.legend = None
            ws.add_chart(bar, f"F{day_top}")

        for col, width in {"A": 34, "B": 12, "C": 10, "D": 8}.items():
            ws.column_dimensions[col].width = width

        # ── Detail sheet: one row per task ──
        ws2 = wb.create_sheet("Detail")
        _head(ws2, 1, ["#", "Date", "Task", "Duration", "Hours"])
        for i, (t, secs) in enumerate(zip(tasks, durations), 1):
            r = i + 1
            ws2.cell(row=r, column=1, value=i).alignment = centered
            ws2.cell(row=r, column=2, value=t.wall_start.strftime("%Y-%m-%d"))
            _text(ws2, r, 3, t.name)
            ws2.cell(row=r, column=4, value=self._fmt_time(secs))
            c = ws2.cell(row=r, column=5, value=round(secs / 3600, 2))
            c.number_format = "0.00"
            for col in range(1, 6):
                ws2.cell(row=r, column=col).border = thin
        total_row = len(tasks) + 2
        ws2.cell(row=total_row, column=3, value="TOTAL").font = bold
        ws2.cell(row=total_row, column=4, value=self._fmt_time(total)).font = bold
        c = ws2.cell(row=total_row, column=5, value=round(total / 3600, 2))
        c.font = bold
        c.number_format = "0.00"
        ws2.auto_filter.ref = f"A1:E{len(tasks) + 1}"
        ws2.freeze_panes = "A2"
        for col, width in {"A": 6, "B": 12, "C": 52, "D": 12, "E": 9}.items():
            ws2.column_dimensions[col].width = width

        return wb

    # ── Report page (for Share a link) ───────────────────────────────────

    def _report_html(self, d_from, d_to, tasks, durations, total, xlsx_b64, xlsx_name) -> str:
        """Self-contained report page: no network, no server, safe to send on."""
        top = self._top_tasks(tasks, durations, total, 10)
        days = self._export_by_day(tasks, durations, d_from, d_to)
        by_day = {d: s for d, s in days if s > 0}
        # A period can hold tasks that all round to zero (started and stopped in
        # the same second). Falling back to the day each task began keeps max()
        # and the average off empty sequences.
        if not by_day:
            by_day = {t.wall_start.date(): 0.0 for t in tasks}
        busiest = max(by_day.items(), key=lambda kv: kv[1])
        avg = total / len(by_day)
        e = _esc

        # A lone day charted by itself is a full-height rectangle that says nothing.
        # Widen the window to its week so the day reads as one day among seven.
        chart_days, ctx_note = days, ""
        if (d_to - d_from).days < 6:
            c_from = d_from - timedelta(days=d_from.weekday())
            c_to = c_from + timedelta(days=6)
            ctx_tasks = self._collect_tasks(c_from, c_to)
            active = self._active_seconds()
            chart_days = self._export_by_day(
                ctx_tasks, [self._task_secs(t, active) for t in ctx_tasks], c_from, c_to)
            ctx_note = ("Shown inside the week of "
                        f"{c_from.strftime('%b %-d')} – {c_to.strftime('%b %-d')}. "
                        "Muted bars fall outside this report.")

        # One measure, one hue: bars scale against the biggest task, so a long tail
        # of near-unique names stays readable instead of collapsing into "Other".
        widest = max((s for _, s, _ in top), default=1.0) or 1.0
        top_html = "".join(
            f'<li><div class="tth"><span class="ttn">{e(n)}</span>'
            f'<span class="ttv num">{self._fmt_time(s)}<b>{sh * 100:.1f}%</b></span></div>'
            f'<div class="ttb"><i data-w="{s / widest * 100:.1f}%"></i></div></li>'
            for n, s, sh in top
        )

        hero = f'<span id="hero" data-secs="{int(total)}">{self._fmt_time(total)}</span>'
        longest = max(durations)
        if d_from == d_to:
            # On a single day "avg/day" and "busiest day" only repeat the total.
            # No wall-clock times either: hours added by hand make a start time
            # disagree with the duration beside it, which only invites questions.
            tiles = [
                ("Tracked", hero, f"{total / 3600:.2f} h", True),
                ("Tasks", str(len(tasks)), "logged", False),
                ("Longest task", self._fmt_time(longest), f"{longest / 3600:.2f} h", False),
                ("Average task", self._fmt_time(total / len(tasks)),
                 f"{total / len(tasks) / 3600:.2f} h", False),
            ]
        else:
            tiles = [
                ("Tracked", hero, f"{total / 3600:.2f} h", True),
                ("Days worked", str(len(by_day)), f"of {len(days)} in period", False),
                ("Tasks", str(len(tasks)), "logged", False),
                ("Avg / day", self._fmt_time(avg), f"{avg / 3600:.2f} h", False),
                ("Longest task", self._fmt_time(longest), f"{longest / 3600:.2f} h", False),
                ("Busiest day", busiest[0].strftime("%a %-d %b"), self._fmt_time(busiest[1]), False),
            ]

        # No day timeline. A task's wall span includes any pause taken inside it, and
        # state.json keeps only a total of paused seconds, never when they happened.
        # Drawing blocks from wall_start→wall_end therefore claims continuous work:
        # on 2 Mar that would assert 26:16 of work against 12:23 actually tracked.
        # Truthful timelines need per-pause timestamps recorded at pause/resume time.
        tiles_html = "".join(
            f'<div class="tile{" hero" if hero else ""}"><div class="k">{e(k)}</div>'
            f'<div class="v num">{v}</div><div class="n num">{e(n)}</div></div>'
            for k, v, n, hero in tiles
        )

        peak = max((s for _, s in chart_days), default=0.0) or 1.0
        bars = []
        for d, s in chart_days:
            inside = d_from <= d <= d_to
            cls = "bar" + ("" if inside else " ctx") + ("" if s else " zero")
            bars.append(
                f'<div class="{cls}" data-tip="{d.strftime("%a %-d %b")} · '
                f'{self._fmt_time(s)}{"" if inside else " · outside report"}">'
                f'<i data-h="{max(2.0, s / peak * 100):.1f}%"></i></div>'
            )
        bars = "".join(bars)

        wide = len(chart_days) > 10
        step = max(1, len(chart_days) // 12)
        axis = "".join(
            f'<span class="num{"" if not (d_from <= d <= d_to) else " on"}">'
            f'{(d.strftime("%-d") if i % step == 0 else "") if wide else d.strftime("%a %-d")}'
            f"</span>"
            for i, (d, _) in enumerate(chart_days)
        )

        rows = "".join(
            f"<tr><td class=\"n num\">{i}</td><td class=\"num\">{t.wall_start.strftime('%Y-%m-%d')}</td>"
            f"<td class=\"tk\">{e(t.name)}</td><td class=\"n num\">{self._fmt_time(s)}</td>"
            f"<td class=\"n num\">{s / 3600:.2f}</td></tr>"
            for i, (t, s) in enumerate(zip(tasks, durations), 1)
        )

        period = self._period_label(self._export_period)
        title = f"Time Report · {self._project or 'Timex'} · {period}"
        return (
            f'<!doctype html><html lang="en"><head><meta charset="utf-8">'
            f'<meta name="viewport" content="width=device-width,initial-scale=1">'
            f"<title>{e(title)}</title><style>{self._REPORT_CSS}</style></head>"
            f'<body><div id="tip"></div><div class="wrap">'
            f'<div class="brand">⏱ Timex</div><h1>Report on Hours</h1>'
            f'<div class="sub">{e(self._report_author())} · {e(self._project or "Timex")} · '
            f"{e(period)}</div>"
            f'<div class="rule"></div><div class="tiles">{tiles_html}</div>'
            f'<div class="rule"></div><h2>Longest tasks</h2>'
            f'<div class="note">The {len(top)} single tasks that took the most time, '
            f"of {len(tasks)} logged. Percentages are of tracked time.</div>"
            f'<ul class="tt">{top_html}</ul>'
            f'<div class="rule"></div><h2>Hours per day</h2>'
            f'<div class="note">{e(ctx_note) if ctx_note else "One bar per calendar day."}</div>'
            f'<div class="chart"><div class="peak num">peak {self._fmt_time(peak)}</div>'
            f'<div class="bars">{bars}</div></div><div class="xaxis">{axis}</div>'
            f'<div class="rule"></div><h2>Detail</h2>'
            f'<div class="note">Every tracked task, in order.</div><table><thead><tr>'
            f'<th class="n">#</th><th>Date</th>'
            f'<th class="tk">Task</th><th class="n">Duration</th><th class="n">Hours</th></tr></thead>'
            f"<tbody>{rows}</tbody></table>"
            f"<footer>Generated by Timex · {self._now().strftime('%Y-%m-%d %H:%M')}</footer>"
            f'</div><a class="dl" download="{e(xlsx_name)}" aria-label="Download {e(xlsx_name)}" '
            f'data-tip="Download {e(xlsx_name)}" href="data:application/vnd.'
            f'openxmlformats-officedocument.spreadsheetml.sheet;base64,{xlsx_b64}">'
            f'<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" '
            f'stroke-linecap="round" stroke-linejoin="round" aria-hidden="true">'
            f'<path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/>'
            f'<polyline points="7 10 12 15 17 10"/><line x1="12" x2="12" y1="15" y2="3"/>'
            f"</svg></a>"
            f"<script>{self._REPORT_JS}</script></body></html>"
        )

    # gh is how we push, and a Finder-launched app gets a bare PATH, so look in
    # the places Homebrew and the installer actually put it.
    _GH_CANDIDATES = ("/opt/homebrew/bin/gh", "/usr/local/bin/gh", "/usr/bin/gh")
    REPORTS_REPO = "halinskiy/timex-reports"
    REPORTS_URL = "https://halinskiy.github.io/timex-reports"

    @classmethod
    def _gh_path(cls) -> str | None:
        for p in cls._GH_CANDIDATES:
            if os.path.isfile(p) and os.access(p, os.X_OK):
                return p
        return None

    def _publish_report(self) -> None:
        """Put the report on the web and hand back a link to send a client.

        The address carries a random slug and the page asks not to be indexed,
        but the host is public: anyone holding the link can read it.
        """
        gh = self._gh_path()
        if not gh:
            self._toast("Publishing needs the gh CLI — brew install gh", 5)
            return
        got = self._export_gather()
        if not got:
            return
        d_from, d_to = got[0], got[1]
        try:
            wb = self._build_workbook(*got)
        except ImportError:
            self._toast("Export unavailable: openpyxl is missing", 5)
            return
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_name = self._export_filename(d_from, d_to, "xlsx")
        html = self._report_html(*got, base64.b64encode(buf.getvalue()).decode(), xlsx_name)
        html = html.replace(
            '<meta name="viewport"',
            '<meta name="robots" content="noindex,nofollow"><meta name="viewport"', 1)

        slug = secrets.token_hex(6)
        payload = base64.b64encode(html.encode()).decode()
        self._toast("Publishing…", 30)

        def _push() -> None:
            try:
                proc = subprocess.run(
                    [gh, "api", "-X", "PUT",
                     f"repos/{self.REPORTS_REPO}/contents/r/{slug}/index.html",
                     "-f", f"message=Report {d_from} — {d_to}",
                     "-f", f"content={payload}"],
                    capture_output=True, text=True, timeout=60,
                )
            except (OSError, subprocess.TimeoutExpired) as exc:
                self.call_from_thread(self._toast, f"Publish failed: {exc}", 6)
                return
            if proc.returncode != 0:
                err = (proc.stderr or "").strip().splitlines()
                self.call_from_thread(self._toast, f"Publish failed: {err[-1] if err else '?'}", 6)
                return
            url = f"{self.REPORTS_URL}/r/{slug}/"
            try:
                subprocess.run(["pbcopy"], input=url, text=True, timeout=5)
            except (OSError, subprocess.TimeoutExpired):
                pass
            self.call_from_thread(self._toast, f"Link copied — {url}", 12)

        threading.Thread(target=_push, daemon=True).start()


    def _cmd_new(self) -> None:
        """Stop timer, save session to history, start fresh."""
        if self.state == IDLE:
            self._toast("Nothing to save")
            return
        n_tasks = len(self.tasks)
        active = self._active_seconds()
        self._save_history()
        self._project_history_loaded = False
        self.state = IDLE
        self.tasks = []
        self._last_session_tasks = []
        self.session_start = None
        self.paused_at = None
        self.total_paused = timedelta()
        self._final_active = 0.0
        self._view_mode = "timeline"
        self._update_placeholder()
        self._mark_dirty()
        self._save_state()
        self._toast(f"Saved {n_tasks} tasks, {self._fmt_time(active)}")

    def _cmd_clear(self) -> None:
        self.state = IDLE
        self.tasks = []
        self._last_session_tasks = []
        self.session_start = None
        self.paused_at = None
        self.total_paused = timedelta()
        self._final_active = 0.0
        self._update_placeholder()
        self._mark_dirty()
        self._save_state()
        self._toast("Cleared")

    def _cmd_add_time(self, raw: str) -> None:
        if self.state == IDLE:
            self._toast("Start the timer first")
            return

        # Parse: /add 10min, /add 10 min, /add 1h, /add 30s, /add 1h30m
        arg = raw[4:].strip()
        if not arg:
            self._toast("Usage: /add 10min, /add 1h, /add 30s")
            return

        total = self._parse_duration(arg)
        if total <= 0:
            self._toast("Usage: /add 10min, /add 1h, /add 30s")
            return

        # Adding time = reducing total_paused
        self.total_paused -= timedelta(seconds=total)

        # Format confirmation
        h, rem = divmod(int(total), 3600)
        m, s = divmod(rem, 60)
        parts_str = []
        if h: parts_str.append(f"{h}h")
        if m: parts_str.append(f"{m}m")
        if s: parts_str.append(f"{s}s")
        self._toast(f"Added {' '.join(parts_str)}")
        self._mark_dirty()
        self._save_state()

    def _cmd_remove_time(self, raw: str) -> None:
        if self.state == IDLE:
            self._toast("Start the timer first")
            return

        arg = raw[7:].strip()
        if not arg:
            self._toast("Usage: /remove 10min, /remove 1h, /remove 30s")
            return

        total = self._parse_duration(arg)
        if total <= 0:
            self._toast("Usage: /remove 10min, /remove 1h, /remove 30s")
            return

        # Removing time = increasing total_paused. Cap it at what is actually on
        # the clock: overshooting used to be masked by max(0, ...) in the display
        # and then silently eaten from the next hours worked.
        available = self._active_seconds()
        if total > available:
            self._toast(f"Only {self._fmt_time(available)} on the clock")
            return
        self.total_paused += timedelta(seconds=total)

        h, rem = divmod(int(total), 3600)
        m, s = divmod(rem, 60)
        parts_str = []
        if h: parts_str.append(f"{h}h")
        if m: parts_str.append(f"{m}m")
        if s: parts_str.append(f"{s}s")
        self._toast(f"Removed {' '.join(parts_str)}")
        self._mark_dirty()
        self._save_state()

    def _cmd_help(self) -> None:
        self._enter_view("help", "  /back to return")

    def _render_help(self) -> None:
        commands = [
            ("/start", "Start the timer"),
            ("/pause", "Pause the timer"),
            ("/resume", "Resume paused timer"),
            ("/new", "Stop timer, save session, start fresh day"),
            ("/add <time>", "Add time manually (e.g. /add 10m, /add 1h)"),
            ("/remove <time>", "Remove time (e.g. /remove 10m, /remove 1h)"),
            ("/edit", "Edit task names in timeline"),
            ("/date", "Browse past sessions by date"),
            ("/stats", "Weekly and monthly statistics"),
            ("/export", "Report a period: visual .html page or .xlsx"),
            ("/clear", "Discard current session without saving"),
            ("/notification", "Set reminder interval"),
            ("/reset", "Reset session (discard without saving)"),
            ("/project", "Switch project"),
            ("/help", "Show this help"),
            ("/back", "Return to previous view"),
        ]
        rows = []
        for i, (cmd, desc) in enumerate(commands):
            if i > 0:
                rows.append(Text.from_markup(f"[{SEPARATOR}]{'─' * 50}[/]"))
            rows.append(Text.from_markup(f"[bold {self._accent}]{cmd}[/]"))
            rows.append(Text.from_markup(f"[{DIM}]{desc}[/]"))

        self.query_one("#history", Static).update(Group(*rows))


    def _cmd_notification(self) -> None:
        self._enter_view("notification", "  Enter number or custom interval \u2022 /back to return")

    def _render_notification(self) -> None:

        rows = []

        # Current setting
        cur = self._reminder_interval
        if cur == 0:
            cur_str = "off"
        elif cur >= 3600:
            cur_str = f"every {cur // 3600}h"
        else:
            cur_str = f"every {cur // 60}m"
        rows.append(Text.from_markup(
            f"[bold {TEXT_COLOR}]Current: {cur_str}[/]"
        ))
        rows.append(Text.from_markup(f"[{SEPARATOR}]{'─' * 50}[/]"))

        # Option 1: Off
        marker = f" [{self._accent}]\u2022[/]" if cur == 0 else ""
        rows.append(Text.from_markup(
            f"[bold {self._accent}]1.[/] [{TEXT_COLOR}]Off[/]{marker}"
        ))

        presets = [
            (10 * 60, "10m"),
            (15 * 60, "15m"),
            (20 * 60, "20m"),
            (30 * 60, "30m"),
            (45 * 60, "45m"),
            (60 * 60, "1h"),
            (2 * 3600, "2h"),
        ]

        for i, (secs, label) in enumerate(presets, start=2):
            rows.append(Text.from_markup(f"[{SEPARATOR}]{'─' * 50}[/]"))
            marker = f" [{self._accent}]\u2022[/]" if secs == cur else ""
            rows.append(Text.from_markup(
                f"[bold {self._accent}]{i}.[/] [{TEXT_COLOR}]Every {label}[/]{marker}"
            ))

        rows.append(Text(""))
        rows.append(Text.from_markup(
            f"  [{DIM}]Or type custom interval (e.g. 25m, 1h30m)[/]"
        ))

        self.query_one("#history", Static).update(Group(*rows))

    def _select_notification(self, raw: str) -> None:
        presets = [10*60, 15*60, 20*60, 30*60, 45*60, 60*60, 2*3600]

        if raw.isdigit():
            num = int(raw)
            if num == 1:
                self._reminder_interval = 0
                self._save_notification(0)
                self._leave_view("Reminders: off")
                return
            if 2 <= num <= len(presets) + 1:
                secs = presets[num - 2]
                self._reminder_interval = secs
                self._save_notification(secs)
                label = f"{secs // 3600}h" if secs >= 3600 else f"{secs // 60}m"
                self._leave_view(f"Reminders: every {label}")
                return
            self._toast(f"Enter 1\u2013{len(presets) + 1}")
            return

        # Parse custom interval
        total = int(self._parse_duration(raw))
        if not total:
            self._toast("Usage: 25m, 1h30m, 45min")
            return
        if total < 60:
            self._toast("Minimum interval: 1 minute")
            return
        self._reminder_interval = total
        self._save_notification(total)
        h, rem = divmod(total, 3600)
        m = rem // 60
        label = ""
        if h: label += f"{h}h"
        if m: label += f"{m}m"
        self._leave_view(f"Reminders: every {label}")

    def _save_notification(self, secs: int) -> None:
        self._save_config("reminder_interval", secs)

    # ── Color ─────────────────────────────────────────────────────────


    def _apply_color(self, hex_val: str) -> None:
        self._accent = hex_val
        self._accent_hex = hex_val.lstrip("#").upper()
        self._save_color(hex_val)
        # Update CSS focus border dynamically
        self.query_one("#task-input").styles.border = ("tall", hex_val)
        self._leave_view()

    def _save_color(self, hex_val: str) -> None:
        self._save_config("accent_color", hex_val)

    def _edit_tasks(self) -> list[TaskEntry]:
        """Get the task list used in edit mode."""
        return self.tasks if self.tasks else self._last_session_tasks

    def _cmd_edit(self) -> None:
        if self._view_mode == "date_sessions":
            self._cmd_edit_sessions()
            return
        if self._view_mode == "project":
            self._cmd_project_edit()
            return
        tasks = self._edit_tasks()
        if not tasks:
            self._toast("No tasks to edit")
            return
        self._view_mode = "edit"
        self._edit_index = len(tasks) - 1
        self._editing_task = None
        self._render_history()
        self.call_after_refresh(self._scroll_to_edit_selection)
        inp = self.query_one("#task-input", HistoryInput)
        inp.placeholder = "  \u2191/\u2193 to select \u2022 Enter to rename \u2022 /back to return"

    def _render_edit(self) -> None:

        tasks = self._edit_tasks()
        if not tasks:
            return
        active = self._active_seconds() if self.tasks else None

        rows = []
        for i, task in enumerate(tasks):
            is_current = self.tasks and i == len(tasks) - 1 and task.active_end is None
            dur = task.format_duration(active if is_current else None)
            time_str = task.format_start()
            selected = (i == self._edit_index)

            if i > 0:
                rows.append(Text.from_markup(f"[{SEPARATOR}]{'─' * 50}[/]"))

            header = self._space_between(f"[{DIM}]{time_str}[/]", f"[#888888]{dur}[/]")
            rows.append(header)

            if selected:
                rows.append(Text.from_markup(f"[bold {self._accent}]\u25ba {task.name}[/]"))
            else:
                rows.append(Text.from_markup(f"[{TEXT_COLOR}]  {task.name}[/]"))

        self.query_one("#history", Static).update(Group(*rows))

    def _edit_move(self, direction: int) -> None:
        tasks = self._edit_tasks()
        if not tasks:
            return
        self._edit_index = max(0, min(len(tasks) - 1, self._edit_index + direction))
        self._render_edit()
        self.call_after_refresh(self._scroll_to_edit_selection)

    def _scroll_to_edit_selection(self) -> None:
        """Scroll to make the selected edit task visible after layout."""
        tasks = self._edit_tasks()
        if not tasks:
            return
        scroll = self.query_one("#history-scroll", VerticalScroll)
        idx = self._edit_index
        total = len(tasks)
        if idx <= 0:
            scroll.scroll_home(animate=False)
        elif idx >= total - 1:
            scroll.scroll_end(animate=False)
        else:
            ratio = idx / max(1, total - 1)
            scroll.scroll_to(y=int(ratio * scroll.max_scroll_y), animate=False)

    def _edit_start_rename(self) -> None:
        tasks = self._edit_tasks()
        if not tasks or self._edit_index >= len(tasks):
            return
        self._editing_task = self._edit_index
        inp = self.query_one("#task-input", HistoryInput)
        inp.value = tasks[self._edit_index].name
        inp.cursor_position = len(inp.value)
        inp.placeholder = "  Enter new name \u2022 empty to delete task"

    def _submit_edit(self, raw: str) -> None:
        if self._editing_task is not None:
            tasks = self._edit_tasks()
            idx = self._editing_task

            if raw:
                # Rename
                if 0 <= idx < len(tasks):
                    tasks[idx].name = raw
                    self._save_state()
                    self._toast("Task renamed")
            else:
                # Empty name — delete task and subtract its time
                if 0 <= idx < len(tasks):
                    self._delete_task(tasks, idx)

            self._editing_task = None
            # If no tasks left, exit edit mode
            if not self._edit_tasks():
                self._leave_view()
                return
            self._edit_index = min(self._edit_index, len(self._edit_tasks()) - 1)
            self._render_edit()
            inp = self.query_one("#task-input", HistoryInput)
            inp.placeholder = "  \u2191/\u2193 to select \u2022 Enter to rename \u2022 /back to return"
            return
        self._toast("Press Enter on empty input to rename selected task")

    def _delete_task(self, tasks: list[TaskEntry], idx: int) -> None:
        """Delete a task and subtract its duration from the session."""
        task = tasks[idx]
        is_current = (task.active_end is None)
        active = self._active_seconds()

        # Calculate task duration
        if is_current:
            duration = active - task.active_start
        else:
            duration = (task.active_end or task.active_start) - task.active_start

        if duration <= 0:
            tasks.pop(idx)
            self._save_state()
            self._toast("Task deleted")
            return

        # Shift all subsequent tasks' active times down by duration
        for t in tasks[idx + 1:]:
            t.active_start -= duration
            if t.active_end is not None:
                t.active_end -= duration

        # Remove the task
        tasks.pop(idx)

        # If deleted task was the last (current) one, make previous task current
        if is_current and tasks and tasks[-1].active_end is not None:
            tasks[-1].active_end = None
            tasks[-1].wall_end = None

        # Subtract duration: increase total_paused or reduce final_active
        if self.state == IDLE:
            self._final_active = max(0.0, self._final_active - duration)
        else:
            self.total_paused += timedelta(seconds=duration)

        self._save_state()
        self._toast(f"Task deleted  (\u2212{self._fmt_time(duration)})")

    def _cmd_date(self) -> None:
        self._enter_view("dates", "  Enter number to view date \u2022 /back to return")

    # ── Project ────────────────────────────────────────────────────────

    def _cmd_project(self) -> None:
        self._enter_view("project", "  Enter number or type new project name \u2022 /back to return")

    def _render_project(self) -> None:
        rows = []

        # Current project
        cur = self._project or "No project"
        rows.append(Text.from_markup(
            f"[bold {TEXT_COLOR}]Current: {cur}[/]"
        ))
        rows.append(Text.from_markup(f"[{SEPARATOR}]{'─' * 50}[/]"))

        # List all projects
        projects = []
        if PROJECTS_DIR.exists():
            projects = sorted(
                [d.name for d in PROJECTS_DIR.iterdir() if d.is_dir()],
                key=str.lower,
            )

        if not projects:
            rows.append(Text.from_markup(
                f"  [{DIM}]No projects yet \u2014 type a name to create one[/]"
            ))
        else:
            for i, name in enumerate(projects, start=1):
                if i > 1:
                    rows.append(Text.from_markup(f"[{SEPARATOR}]{'─' * 50}[/]"))
                # Read project state to show status
                pstate, ptime = self._read_project_status(name)
                if pstate == RUNNING:
                    status = f"[bold {self._accent}]\u25cf REC     {ptime}[/]"
                elif pstate == PAUSED:
                    status = f"[bold #888888]\u275a\u275a PAUSED  {ptime}[/]"
                else:
                    status = f"[{DIM}]\u25cb IDLE[/]"
                marker = f" [{self._accent}]\u2022[/]" if name == self._project else ""
                rows.append(self._space_between(
                    f"[bold {self._accent}]{i}.[/] [{TEXT_COLOR}]{name}[/]{marker}",
                    status,
                ))

        rows.append(Text(""))
        rows.append(Text.from_markup(
            f"  [{DIM}]Type a name to create new project[/]"
        ))

        self.query_one("#history", Static).update(Group(*rows))

    def _read_project_status(self, name: str) -> tuple[str, str]:
        """Read a project's state.json and return (state, formatted_time)."""
        sf = PROJECTS_DIR / name / "state.json"
        try:
            if not sf.exists():
                return IDLE, ""
            data = json.loads(sf.read_text())
            state = data.get("state", IDLE)
            # Calculate active seconds from saved data
            total_paused_secs = data.get("total_paused_secs", 0.0)
            session_start_str = data.get("session_start")
            if not session_start_str:
                return state, self._fmt_time(data.get("final_active", 0.0))
            session_start = datetime.fromisoformat(session_start_str)
            now = self._now()
            if state == RUNNING:
                elapsed = (now - session_start).total_seconds() - total_paused_secs
            elif state == PAUSED:
                paused_at_str = data.get("paused_at")
                if paused_at_str:
                    elapsed = (datetime.fromisoformat(paused_at_str) - session_start).total_seconds() - total_paused_secs
                else:
                    elapsed = 0.0
            else:
                return IDLE, self._fmt_time(data.get("final_active", 0.0))
            return state, self._fmt_time(max(0.0, elapsed))
        except (OSError, json.JSONDecodeError, KeyError, ValueError):
            return IDLE, ""

    def _select_project(self, raw: str) -> None:
        projects = []
        if PROJECTS_DIR.exists():
            projects = sorted(
                [d.name for d in PROJECTS_DIR.iterdir() if d.is_dir()],
                key=str.lower,
            )

        if raw.isdigit():
            num = int(raw)
            if 1 <= num <= len(projects):
                self._switch_project(projects[num - 1])
                return
            self._toast(f"Enter 1\u2013{len(projects)}" if projects else "No projects yet")
            return

        # Text input = create new project
        name = raw.strip()
        if not name:
            return
        if name.startswith("/"):
            return  # handled by command routing above
        self._switch_project(name)

    def _switch_project(self, name: str) -> None:
        """Switch to a project (create dir if needed). Keeps current state as-is."""
        if name == self._project:
            self._leave_view()
            return
        # Stop watch before switching (watch is per-project)
        # Save current project state (RUNNING stays RUNNING)
        self._save_state()

        # Save active project
        self._project = name
        try:
            STATE_DIR.mkdir(parents=True, exist_ok=True)
            ACTIVE_PROJECT_FILE.write_text(name)
        except OSError:
            pass

        # Create project dir
        pdir = self._project_dir()
        pdir.mkdir(parents=True, exist_ok=True)

        # Reset state and load new project
        self.state = IDLE
        self.tasks = []
        self.session_start = None
        self.paused_at = None
        self.total_paused = timedelta()
        self._final_active = 0.0
        self._project_history_loaded = False
        self._last_session_tasks = []
        self._last_saved_at = ""

        self._load_state(preserve_running=True)
        self._leave_view(f"Switched to {name}")

    # ── Project Edit ──────────────────────────────────────────────────

    def _cmd_project_edit(self) -> None:
        projects = []
        if PROJECTS_DIR.exists():
            projects = sorted(
                [d.name for d in PROJECTS_DIR.iterdir() if d.is_dir()],
                key=str.lower,
            )
        if not projects:
            self._toast("No projects to edit")
            return
        self._project_edit_index = 0
        self._project_editing = None
        self._enter_view("project_edit", "  ↑/↓ to select • Enter to rename • /back")

    def _render_project_edit(self) -> None:
        projects = []
        if PROJECTS_DIR.exists():
            projects = sorted(
                [d.name for d in PROJECTS_DIR.iterdir() if d.is_dir()],
                key=str.lower,
            )
        if not projects:
            self._leave_view("No projects")
            return

        rows = []
        for i, name in enumerate(projects):
            if i > 0:
                rows.append(Text.from_markup(f"[{SEPARATOR}]{'─' * 50}[/]"))
            pstate, ptime = self._read_project_status(name)
            if pstate == RUNNING:
                status = f"[bold {self._accent}]● REC     {ptime}[/]"
            elif pstate == PAUSED:
                status = f"[bold #888888]❚❚ PAUSED  {ptime}[/]"
            else:
                status = f"[{DIM}]○ IDLE[/]"
            selected = (i == self._project_edit_index)
            marker = f" [{self._accent}]•[/]" if name == self._project else ""
            if selected:
                rows.append(self._space_between(
                    f"[bold {self._accent}]▸ {name}[/]{marker}",
                    status,
                ))
            else:
                rows.append(self._space_between(
                    f"[{TEXT_COLOR}]  {name}[/]{marker}",
                    status,
                ))

        self.query_one("#history", Static).update(Group(*rows))

    def _project_edit_move(self, direction: int) -> None:
        projects = []
        if PROJECTS_DIR.exists():
            projects = sorted(
                [d.name for d in PROJECTS_DIR.iterdir() if d.is_dir()],
                key=str.lower,
            )
        if not projects:
            return
        self._project_edit_index = max(0, min(len(projects) - 1, self._project_edit_index + direction))
        self._render_project_edit()

    def _project_edit_start_rename(self) -> None:
        projects = []
        if PROJECTS_DIR.exists():
            projects = sorted(
                [d.name for d in PROJECTS_DIR.iterdir() if d.is_dir()],
                key=str.lower,
            )
        if not projects or self._project_edit_index >= len(projects):
            return
        self._project_editing = self._project_edit_index
        inp = self.query_one("#task-input", HistoryInput)
        inp.value = projects[self._project_edit_index]
        inp.cursor_position = len(inp.value)
        inp.placeholder = "  Enter new name • empty to delete"

    def _select_project_edit(self, raw: str) -> None:
        projects = []
        if PROJECTS_DIR.exists():
            projects = sorted(
                [d.name for d in PROJECTS_DIR.iterdir() if d.is_dir()],
                key=str.lower,
            )

        if self._project_editing is None:
            # Not in rename mode — start rename on Enter with text
            return

        idx = self._project_editing
        if idx >= len(projects):
            self._project_editing = None
            return

        old_name = projects[idx]

        if raw:
            # Rename
            new_name = raw.strip()
            if not new_name or new_name == old_name:
                self._project_editing = None
                self._render_project_edit()
                inp = self.query_one("#task-input", HistoryInput)
                inp.placeholder = "  ↑/↓ to select • Enter to rename • /back"
                return
            old_path = PROJECTS_DIR / old_name
            new_path = PROJECTS_DIR / new_name
            if new_path.exists():
                self._toast(f"Project '{new_name}' already exists")
                return
            try:
                old_path.rename(new_path)
            except OSError as e:
                self._toast(f"Rename failed: {e}")
                return
            # Update active project if renamed
            if self._project == old_name:
                self._project = new_name
                try:
                    ACTIVE_PROJECT_FILE.write_text(new_name)
                except OSError:
                    pass
            self._project_editing = None
            self._render_project_edit()
            inp = self.query_one("#task-input", HistoryInput)
            inp.placeholder = "  ↑/↓ to select • Enter to rename • /back"
            self._toast(f"Renamed → {new_name}")
        else:
            # Empty name — delete
            self._project_to_delete = old_name
            self._project_editing = None
            self._enter_view("confirm_delete_project", "  y to confirm, n to cancel")

    # ── Confirm Delete Project ────────────────────────────────────────

    def _render_confirm_delete_project(self) -> None:
        name = self._project_to_delete or "?"
        rows = [
            Text(""),
            Text.from_markup(f"[bold {self._accent}]Delete project '{name}'?[/]"),
            Text(""),
            Text.from_markup(f"[{DIM}]All data will be lost.[/]"),
        ]
        rows += [
            Text.from_markup(f"[{SEPARATOR}]{'─' * 50}[/]"),
            Text.from_markup(f"[bold {self._accent}]y.[/] [{TEXT_COLOR}]Confirm delete[/]"),
            Text.from_markup(f"[{SEPARATOR}]{'─' * 50}[/]"),
            Text.from_markup(f"[bold {self._accent}]n.[/] [{TEXT_COLOR}]Cancel[/]"),
        ]
        self.query_one("#history", Static).update(Group(*rows))

    def _select_confirm_delete_project(self, raw: str) -> None:
        if raw.lower() not in ("y", "yes", "n", "no"):
            return
        if raw.lower() in ("y", "yes"):
            name = self._project_to_delete
            if name:
                target = PROJECTS_DIR / name
                if target.exists():
                    if not self._snapshot("project-delete"):
                        self._toast("Backup failed — project kept", 5)
                        self._project_to_delete = None
                        self._enter_view("project_edit", "  ↑/↓ to select • Enter to rename • /back")
                        return
                    shutil.rmtree(target)
                # If deleted current project, switch to another
                if self._project == name:
                    projects = []
                    if PROJECTS_DIR.exists():
                        projects = sorted(
                            [d.name for d in PROJECTS_DIR.iterdir() if d.is_dir()],
                            key=str.lower,
                        )
                    if projects:
                        self._project = projects[0]
                        try:
                            ACTIVE_PROJECT_FILE.write_text(projects[0])
                        except OSError:
                            pass
                        self._load_state()
                    else:
                        self._project = None
                        try:
                            ACTIVE_PROJECT_FILE.unlink(missing_ok=True)
                        except OSError:
                            pass
                        # Reset to default state dir
                        self.state = IDLE
                        self.tasks = []
                        self._last_session_tasks = []
                        self.session_start = None
                        self.paused_at = None
                        self.total_paused = timedelta()
                        self._final_active = 0.0
                        self._load_state()
            self._project_to_delete = None
            # Go back to project_edit if projects remain
            projects = []
            if PROJECTS_DIR.exists():
                projects = sorted(
                    [d.name for d in PROJECTS_DIR.iterdir() if d.is_dir()],
                    key=str.lower,
                )
            if projects:
                self._project_edit_index = min(self._project_edit_index, len(projects) - 1)
                self._enter_view("project_edit", "  ↑/↓ to select • Enter to rename • /back")
                self._toast(f"Project '{name}' deleted")
            else:
                self._leave_view(f"Project '{name}' deleted")
        else:
            self._project_to_delete = None
            self._enter_view("project_edit", "  ↑/↓ to select • Enter to rename • /back")

    def _cmd_stats(self) -> None:
        self._enter_view("stats", "  /back to return")

    def _render_stats(self) -> None:
        from collections import defaultdict
        from rich.console import Group

        # Collect all sessions: history + current
        history = self._load_history()
        if self.state in (RUNNING, PAUSED) and self.tasks:
            active = self._active_seconds()
            serialized = [self._serialize_task(t) for t in self.tasks]
            if serialized and serialized[-1].get("active_end") is None:
                serialized[-1]["active_end"] = active  # count the task in progress
            current = {
                "date": self.tasks[0].wall_start.strftime("%Y-%m-%d"),
                "total_active": active,
                "tasks": serialized,
            }
            history = history + [current]

        if not history:
            self.query_one("#history", Static).update(
                Text.from_markup(f"\n  [white]No data yet — complete a session first[/]\n")
            )
            return

        today = self._now().date()

        # Group by date → total seconds
        by_date: dict[str, float] = defaultdict(float)
        sessions_by_date: dict[str, int] = defaultdict(int)
        for session in history:
            d = session.get("date", "")
            by_date[d] += session.get("total_active", 0)
            sessions_by_date[d] += 1

        # Compute 7-day and 30-day stats
        def _period_stats(days: int) -> tuple[float, float, int]:
            total = 0.0
            count = 0
            for i in range(days):
                d = (today - timedelta(days=i)).strftime("%Y-%m-%d")
                total += by_date.get(d, 0)
                count += sessions_by_date.get(d, 0)
            avg = total / days if days else 0
            return total, avg, count

        total_7, avg_7, sess_7 = _period_stats(7)
        total_30, avg_30, sess_30 = _period_stats(30)

        rows = []

        # ── 7-day summary ──
        rows.append(self._space_between(
            f"[bold {TEXT_COLOR}]Last 7 days[/]",
            f"[bold {self._accent}]{self._fmt_time(total_7)}[/]",
        ))
        rows.append(Text.from_markup(
            f"[{DIM}]Avg/day  {self._fmt_time(avg_7)}  \u00b7  Sessions  {sess_7}[/]"
        ))
        rows.append(Text(""))

        # ── 30-day summary ──
        rows.append(self._space_between(
            f"[bold {TEXT_COLOR}]Last 30 days[/]",
            f"[bold {self._accent}]{self._fmt_time(total_30)}[/]",
        ))
        rows.append(Text.from_markup(
            f"[{DIM}]Avg/day  {self._fmt_time(avg_30)}  \u00b7  Sessions  {sess_30}[/]"
        ))

        # ── Bar chart: last 7 days ──
        rows.append(Text.from_markup(f"[{SEPARATOR}]{'─' * 50}[/]"))

        bar_width = 30
        day_data = []
        for i in range(6, -1, -1):
            d = today - timedelta(days=i)
            secs = by_date.get(d.strftime("%Y-%m-%d"), 0)
            day_data.append((d.strftime("%a"), secs))

        max_secs = max((s for _, s in day_data), default=0)

        for idx, (label, secs) in enumerate(day_data):
            if max_secs > 0 and secs > 0:
                filled = max(1, round(secs / max_secs * bar_width))
            else:
                filled = 0
            empty = bar_width - filled
            bar = f"[{self._accent}]{'█' * filled}[/][#555555]{'░' * empty}[/]"
            time_str = f"[{TEXT_COLOR}]{self._fmt_time(secs)}[/]" if secs > 0 else f"[{DIM}]—[/]"
            if idx > 0:
                rows.append(Text(""))
            rows.append(self._space_between(
                f"[{DIM}]{label}[/]    {bar}",
                time_str,
            ))

        # ── Top tasks (30 days) ──
        # Collect task times for last 30 days only
        task_times_30: dict[str, float] = defaultdict(float)
        for session in history:
            d = session.get("date", "")
            try:
                session_date = datetime.strptime(d, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                continue
            if (today - session_date).days >= 30:
                continue
            for t in session.get("tasks", []):
                start = t.get("active_start", 0) or 0
                end = t.get("active_end") or t.get("active_start", 0) or 0
                dur = max(0, end - start)
                if dur > 0:
                    task_times_30[t.get("name", "Untitled")] += dur

        if task_times_30:
            rows.append(Text.from_markup(f"[{SEPARATOR}]{'─' * 50}[/]"))
            rows.append(Text.from_markup(f"[bold {TEXT_COLOR}]Longest tasks (30 days)[/]"))

            # Biggest tasks, not a breakdown — names are near-unique, so folding the
            # tail into "Other" would just draw one huge meaningless bar.
            ranked = sorted(task_times_30.items(), key=lambda kv: kv[1], reverse=True)[:5]
            widest = max((v for _, v in ranked), default=0.0) or 1.0
            grand = sum(task_times_30.values())

            for name, secs in ranked:
                rows.append(Text(""))
                filled = max(1, round(secs / widest * 22))
                bar = f"[{self._accent}]{'█' * filled}[/][{DIMMER}]{'░' * (22 - filled)}[/]"
                pct = secs / grand * 100 if grand else 0.0
                display_name = name if len(name) <= 42 else name[:39] + "..."
                rows.append(Text.from_markup(f"[{TEXT_COLOR}]{display_name}[/]"))
                rows.append(self._space_between(
                    f"{bar}", f"[{DIM}]{self._fmt_time(secs)}[/]  [{TEXT_COLOR}]{pct:4.1f}%[/]",
                ))

        self.query_one("#history", Static).update(Group(*rows))

    def _cmd_back(self) -> None:
        if self._view_mode == "history_detail":
            if self._viewing_sessions:
                # Back to session list
                self._view_mode = "date_sessions"
                self._viewing_tasks = []
                self._render_history()
                inp = self.query_one("#task-input", HistoryInput)
                inp.placeholder = "  Enter number \u2022 /edit to manage \u2022 /back"
            else:
                self._view_mode = "dates"
                self._viewing_tasks = []
                self._viewing_date = ""
                self._render_history()
                inp = self.query_one("#task-input", HistoryInput)
                inp.placeholder = "  Enter number to view date \u2022 /back to return"
        elif self._view_mode == "date_sessions":
            self._viewing_sessions = []
            self._view_mode = "dates"
            self._render_history()
            inp = self.query_one("#task-input", HistoryInput)
            inp.placeholder = "  Enter number to view date \u2022 /back to return"
        elif self._view_mode == "edit_sessions":
            self._editing_session = None
            self._view_mode = "date_sessions"
            self._render_history()
            inp = self.query_one("#task-input", HistoryInput)
            inp.placeholder = "  Enter number \u2022 /edit to manage \u2022 /back"
        elif self._view_mode == "confirm_delete_project":
            self._project_to_delete = None
            self._enter_view("project_edit", "  ↑/↓ to select • Enter to rename • /back")
        elif self._view_mode == "project_edit":
            self._project_editing = None
            self._enter_view("project", "  Enter number or type new project name • /back to return")
        elif self._view_mode in ("dates", "help", "notification", "edit", "color", "stats", "project", "confirm_reset", "export", "update"):
            self._editing_task = None
            self._leave_view()
        else:
            self._toast("Already on Timeline")

    def _select_date(self, num: int) -> None:
        if num < 1 or num > len(self._dates_list):
            self._toast(f"Enter 1\u2013{len(self._dates_list)}")
            return

        date_str = self._dates_list[num - 1]
        history = self._load_history()

        sessions = [s for s in history if s.get("date") == date_str]
        if not sessions:
            self._toast("No data for this date")
            return

        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            nice_date = dt.strftime("%a, %b %d %Y")
        except ValueError:
            nice_date = date_str

        if len(sessions) == 1:
            # Single session → go straight to tasks
            tasks = [self._deserialize_task(td) for td in sessions[0].get("tasks", [])]
            if not tasks:
                self._toast("No tasks for this date")
                return
            self._view_mode = "history_detail"
            self._viewing_tasks = tasks
            self._viewing_date = nice_date
            self._viewing_date_str = date_str
            self._render_history()
            inp = self.query_one("#task-input", HistoryInput)
            inp.placeholder = "  /edit to manage \u2022 /back to return"
        else:
            # Multiple sessions → show session list
            self._viewing_sessions = sessions
            self._viewing_date = nice_date
            self._viewing_date_str = date_str
            self._view_mode = "date_sessions"
            self._render_history()
            inp = self.query_one("#task-input", HistoryInput)
            inp.placeholder = "  Enter number \u2022 /edit to manage \u2022 /back"

    # ── Helpers ──────────────────────────────────────────────────────────

    def _toast(self, message: str, seconds: float = 3) -> None:
        """Show a message above the input field, then auto-clear."""
        toast = self.query_one("#toast-bar", Static)
        toast.update(Text.from_markup(
            f"[bold #171717 on {self._accent}] {message} [/]"
        ))
        self.set_timer(seconds, lambda: toast.update(""))

    def _update_placeholder(self) -> None:
        inp = self.query_one("#task-input", HistoryInput)
        if self.state == RUNNING:
            inp.placeholder = "  What are you working on?"
        elif self.state == PAUSED:
            inp.placeholder = "  Timer paused \u2014 /resume to continue"
        else:
            inp.placeholder = "  Type a task to start tracking"

    # ── Persistence ──────────────────────────────────────────────────────

    @staticmethod
    def _serialize_task(t: TaskEntry) -> dict:
        d = {
            "name": t.name,
            "wall_start": t.wall_start.isoformat(),
            "active_start": t.active_start,
            "active_end": t.active_end,
            "wall_end": t.wall_end.isoformat() if t.wall_end else None,
        }
        return d

    @staticmethod
    def _deserialize_task(d: dict) -> TaskEntry:
        return TaskEntry(
            name=d["name"],
            wall_start=datetime.fromisoformat(d["wall_start"]),
            active_start=d["active_start"],
            active_end=d.get("active_end"),
            wall_end=datetime.fromisoformat(d["wall_end"]) if d.get("wall_end") else None,
        )

    def _save_state(self) -> None:
        """Persist the session — a blind overwrite of whatever is on disk.

        Callers must absorb external writes BEFORE mutating state (see
        _on_submit), never here: by this point the user's change is already in
        memory, and reloading would throw it away.
        """
        try:
            STATE_DIR.mkdir(parents=True, exist_ok=True)
            saved_at = self._now().isoformat()
            data = {
                "state": self.state,
                "tasks": [self._serialize_task(t) for t in self.tasks],
                "last_session_tasks": [self._serialize_task(t) for t in self._last_session_tasks],
                "session_start": self.session_start.isoformat() if self.session_start else None,
                "paused_at": self.paused_at.isoformat() if self.paused_at else None,
                "total_paused_secs": self.total_paused.total_seconds(),
                "final_active": self._final_active,
                "saved_at": saved_at,
            }
            sf = self._state_file()
            sf.parent.mkdir(parents=True, exist_ok=True)
            tmp = sf.with_suffix(".tmp")
            tmp.write_text(json.dumps(data, indent=2))
            tmp.replace(sf)
            self._last_saved_at = saved_at
        except OSError:
            pass

    def _build_history_entry(self) -> dict | None:
        """Build history entry from current session WITHOUT modifying state."""
        if not self.tasks:
            return None
        active = self._active_seconds()
        now = self._now()
        # Finalize last task timestamps in-place
        if self.tasks and self.tasks[-1].active_end is None:
            self.tasks[-1].active_end = active
            self.tasks[-1].wall_end = now
        return {
            "date": self.tasks[0].wall_start.strftime("%Y-%m-%d"),
            "session_start": self.session_start.isoformat() if self.session_start else None,
            "total_active": active,
            "tasks": [self._serialize_task(t) for t in self.tasks],
        }

    def _append_history(self, entry: dict) -> None:
        """Append a pre-built entry to history.json."""
        try:
            history = self._load_history()
            history.append(entry)
            hf = self._history_file()
            hf.parent.mkdir(parents=True, exist_ok=True)
            tmp = hf.with_suffix(".tmp")
            tmp.write_text(json.dumps(history, indent=2))
            tmp.replace(hf)
            self._invalidate_history_cache()
            self._reload_project_history_secs()
        except OSError:
            pass

    def _save_history(self) -> None:
        """Append current session to history.json."""
        entry = self._build_history_entry()
        if entry:
            self._append_history(entry)

    # ── Backups ──────────────────────────────────────────────────────────

    def _snapshot(self, tag: str, once_per_day: bool = False) -> bool:
        """Copy the projects tree aside before anything can destroy it.

        history.json is the only record a session ever gets, so a delete or a bad
        write is otherwise final. Restoring is a plain folder copy back.
        Returns False if the copy failed, so callers can refuse to destroy.
        """
        try:
            if not PROJECTS_DIR.exists() or not any(PROJECTS_DIR.iterdir()):
                return True  # nothing to lose
            BACKUP_DIR.mkdir(parents=True, exist_ok=True)
            now = self._now()
            if once_per_day:
                today = now.strftime("%Y%m%d")
                for p in BACKUP_DIR.iterdir():
                    if p.is_dir() and p.name.startswith(today) and p.name.endswith(tag):
                        return True  # already snapshotted today
            shutil.copytree(PROJECTS_DIR,
                            BACKUP_DIR / f"{now.strftime('%Y%m%d-%H%M%S')}-{tag}")
            snaps = sorted(p for p in BACKUP_DIR.iterdir() if p.is_dir())
            for old in snaps[:-BACKUP_KEEP]:
                shutil.rmtree(old, ignore_errors=True)
            return True
        except OSError:
            return False  # never take the app down, but never claim success either

    _history_cache: list[dict] | None = None
    _history_cache_mtime: float = 0.0
    _history_cache_path: str = ""

    def _load_history(self) -> list[dict]:
        try:
            hf = self._history_file()
            hf_str = str(hf)
            if not hf.exists():
                return []
            mtime = hf.stat().st_mtime
            if (self._history_cache is not None
                    and mtime == self._history_cache_mtime
                    and hf_str == self._history_cache_path):
                return self._history_cache
            data = json.loads(hf.read_text())
            if not isinstance(data, list):
                return []  # valid JSON of the wrong shape would break every caller
            TimexApp._history_cache = data
            TimexApp._history_cache_mtime = mtime
            TimexApp._history_cache_path = hf_str
            return data
        except (OSError, json.JSONDecodeError):
            pass
        return []

    def _invalidate_history_cache(self) -> None:
        TimexApp._history_cache = None

    def _reload_project_history_secs(self) -> None:
        """Cache total active seconds from project history."""
        history = self._load_history() or []
        self._project_history_secs = sum(s.get("total_active", 0.0) for s in history)
        self._project_history_loaded = True

    def _project_total_seconds(self) -> float:
        """Total active seconds for project: history + current session."""
        if not self._project_history_loaded:
            self._reload_project_history_secs()
        return self._project_history_secs + self._active_seconds()

    def _all_sessions_active_seconds(self) -> float:
        """Sum of active (session) seconds across all projects right now."""
        total = 0.0
        if PROJECTS_DIR.exists():
            for d in PROJECTS_DIR.iterdir():
                if d.is_dir():
                    if d.name == self._project:
                        total += self._active_seconds()
                        continue
                    sf = d / "state.json"
                    if not sf.exists():
                        continue
                    try:
                        data = json.loads(sf.read_text())
                        st = data.get("state", IDLE)
                        if st == IDLE:
                            continue
                        ss_str = data.get("session_start")
                        if not ss_str:
                            continue
                        ss = datetime.fromisoformat(ss_str)
                        tp = timedelta(seconds=data.get("total_paused_secs", 0.0))
                        if st == RUNNING:
                            elapsed = (self._now() - ss) - tp
                        elif st == PAUSED:
                            pa_str = data.get("paused_at")
                            elapsed = (datetime.fromisoformat(pa_str) - ss) - tp if pa_str else (self._now() - ss) - tp
                        else:
                            continue
                        total += max(0.0, elapsed.total_seconds())
                    except (OSError, json.JSONDecodeError, ValueError):
                        pass
        return total

    def _load_state(self, preserve_running: bool = False) -> None:
        try:
            sf = self._state_file()
            if not sf.exists():
                return
            data = json.loads(sf.read_text())
        except (OSError, json.JSONDecodeError, KeyError):
            return

        try:
            saved_state = data.get("state", IDLE)
            self.tasks = [self._deserialize_task(d) for d in data.get("tasks", [])]
            self._last_session_tasks = [self._deserialize_task(d) for d in data.get("last_session_tasks", [])]
            self._final_active = data.get("final_active", 0.0)
            self.total_paused = timedelta(seconds=data.get("total_paused_secs", 0.0))

            session_start_str = data.get("session_start")
            self.session_start = datetime.fromisoformat(session_start_str) if session_start_str else None

            if saved_state in (RUNNING, PAUSED):
                if saved_state == RUNNING and preserve_running:
                    # Project switch — was running moments ago, keep running
                    self.paused_at = None
                    self.state = RUNNING
                elif saved_state == RUNNING:
                    # Cold start — pause at the moment of last save
                    saved_at_str = data.get("saved_at")
                    if saved_at_str:
                        self.paused_at = datetime.fromisoformat(saved_at_str)
                    else:
                        self.paused_at = self._now()
                    self.state = PAUSED
                else:    # Was already paused — keep original paused_at.
                    paused_at_str = data.get("paused_at")
                    self.paused_at = datetime.fromisoformat(paused_at_str) if paused_at_str else self._now()
                    self.state = PAUSED
                self._reset_reminder()

                self._save_state()  # also sets self._last_saved_at
            elif saved_state == IDLE:
                self.state = IDLE
                self._last_saved_at = data.get("saved_at", "")
        except (KeyError, ValueError, TypeError):
            pass

    # ── Reminders ────────────────────────────────────────────────────────

    def _reset_reminder(self) -> None:
        """Reset the reminder countdown (called on state changes)."""
        self._last_reminder = _time.monotonic()

    def _check_reminder(self) -> None:
        """Fire a reminder every REMINDER_INTERVAL seconds while not idle."""
        if self._reminder_interval == 0:
            return
        now = _time.monotonic()
        if self._last_reminder == 0.0:
            self._last_reminder = now
            return
        if now - self._last_reminder >= self._reminder_interval:
            self._last_reminder = now
            self._send_reminder()

    def _send_reminder(self) -> None:
        """Show in-app + macOS system notification."""
        # Guard against double-fire (e.g. after system wake queuing multiple ticks)
        now = _time.monotonic()
        if now - getattr(self, '_last_notify_at', 0.0) < 10.0:
            return
        self._last_notify_at = now

        elapsed_str = self._fmt_time(self._active_seconds())
        current_task = self.tasks[-1].name if self.tasks and self.tasks[-1].active_end is None else None

        if self.state == RUNNING:
            app_msg = f"Still Recording  [{elapsed_str}]"
            sys_msg = "Still Recording"
        elif self.state == PAUSED:
            app_msg = f"Still Paused  [{elapsed_str}]"
            sys_msg = "Still Paused"
        else:
            return

        self._toast(app_msg, 8)
        self._system_notify(sys_msg)

    @staticmethod
    def _system_notify(message: str) -> None:
        """Send macOS notification via Swift helper."""
        helper = Path(__file__).parent / "TimexNotify.app" / "Contents" / "MacOS" / "timex-notify"
        try:
            subprocess.Popen(
                [str(helper), "Timex", message],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
        except OSError:
            pass


    # ── Auto-update ────────────────────────────────────────────────────────


    def action_quit(self) -> None:
        # Auto-pause on exit so no reminders fire while app is closed
        if self.state == RUNNING:
            self.state = PAUSED
            self.paused_at = self._now()
            self._save_state()
        self.exit()


# ── Entry point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import logging
    logging.basicConfig(
        filename=str(Path.home() / ".timex" / "debug.log"),
        level=logging.DEBUG,
        format="%(asctime)s %(levelname)s %(message)s",
    )
    try:
        TimexApp().run()
    except Exception:
        logging.exception("App crashed")
        raise
